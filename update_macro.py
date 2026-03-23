import pandas as pd
import openpyxl
import os
import glob

# =============================================
# 경로 자동 설정 - 건드릴 필요 없어!
# =============================================

# 각자 컴퓨터 사용자 폴더 자동 감지
base = os.path.expanduser("~")

# SharePoint 기본 경로
sharepoint_base = os.path.join(base, "OneDrive - 잡코리아", "전략추진실 - 문서", "4. Macro Analysis - 핵심 선후행 지표")

# KOSIS 파일 폴더 (매월 여기에 넣어줘!)
KOSIS_FOLDER = os.path.join(sharepoint_base, "자동화 툴", "01_KOSIS 데이터")

# macro 파일 경로
MACRO_FILE = os.path.join(sharepoint_base, "연습_✭Macro Analysis.xlsx")

# =============================================
# 지표 매핑: KOSIS 컬럼명 → macro 시트명
# =============================================
SHEET_MAPPING = {
    "빈일자리_상용 (명)":       "빈일자리_상용",
    "빈일자리_임시일용 (명)":   "빈일자리_임시일용",
    "빈일자리율_상용 (%)":      "빈일자리율_상용",
    "빈일자리율_임시일용 (%)":  "빈일자리율_임시일용",
    "채용_상용 (명)":           "채용_상용",
    "채용_임시일용 (명)":       "채용_임시일용",
    "근로자_상용 (명)":         "근로자_상용",
    "근로자_임시일용 (명)":     "근로자_임시일용",
    "입직자_상용 (명)":         "입직자_상용",
    "입직자_임시일용 (명)":     "입직자_임시일용",
}

print("=" * 50)
print("macro 파일 업데이트 시작!")
print("=" * 50)
print(f"\n사용자 경로: {base}")

# =============================================
# 1. KOSIS 파일 자동 탐색 (가장 최신 파일)
# =============================================
print(f"\n[1단계] KOSIS 파일 자동 탐색 중...")
print(f"  → 탐색 폴더: {KOSIS_FOLDER}")

if not os.path.exists(KOSIS_FOLDER):
    print(f"❌ KOSIS 폴더가 없어!")
    print(f"   → SharePoint에 '자동화 툴 > 01_KOSIS 데이터' 폴더를 만들어줘!")
    exit()

kosis_pattern = os.path.join(KOSIS_FOLDER, "산업_규모별_고용_*.xlsx")
kosis_files = glob.glob(kosis_pattern)

if not kosis_files:
    print(f"❌ KOSIS 파일을 찾을 수 없어!")
    print(f"   → '01_KOSIS 데이터' 폴더에 '산업_규모별_고용_*.xlsx' 파일을 넣어줘!")
    exit()

kosis_files.sort()
KOSIS_FILE = kosis_files[-1]
print(f"  → 찾은 파일: {os.path.basename(KOSIS_FILE)}")
if len(kosis_files) > 1:
    print(f"  → 여러 파일 중 가장 최신 파일 사용 (총 {len(kosis_files)}개)")

# =============================================
# 2. KOSIS 파일 읽기
# =============================================
print(f"\n[2단계] KOSIS 파일 읽는 중...")
kosis_df = pd.read_excel(KOSIS_FILE, sheet_name='데이터', header=None)
new_month = str(kosis_df.iloc[0, 2])
print(f"  → 새로 추가할 월: {new_month}")

indicator_row = kosis_df.iloc[1].tolist()
kosis_df[0] = kosis_df[0].ffill()

# =============================================
# 3. macro 파일 열기
# =============================================
print(f"\n[3단계] macro 파일 여는 중...")
print(f"  → {MACRO_FILE}")

if not os.path.exists(MACRO_FILE):
    print(f"❌ macro 파일을 찾을 수 없어!")
    print(f"   → 경로 확인: {MACRO_FILE}")
    exit()

wb = openpyxl.load_workbook(MACRO_FILE)
updated_sheets = []
skipped_sheets = []

# =============================================
# 4. 각 시트 업데이트
# =============================================
print(f"\n[4단계] 시트 업데이트 중...")
for kosis_col_name, sheet_name in SHEET_MAPPING.items():
    if sheet_name not in wb.sheetnames:
        print(f"  ⚠ {sheet_name}: 시트를 찾을 수 없어!")
        skipped_sheets.append(sheet_name)
        continue

    ws = wb[sheet_name]
    header_row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    indicator_header = list(ws.iter_rows(min_row=3, max_row=3, values_only=True))[0]

    # 해당 월 컬럼 인덱스 찾기
    col_idx = None
    for i, val in enumerate(header_row):
        if str(val) == new_month:
            col_idx = i + 1
            break

    # 없으면 새 컬럼 추가
    if col_idx is None:
        last_date_col = None
        for i, val in enumerate(header_row):
            if val and str(val).startswith('20'):
                last_date_col = i + 1
        if last_date_col is None:
            print(f"  ⚠ {sheet_name}: 날짜 컬럼을 찾을 수 없어!")
            skipped_sheets.append(sheet_name)
            continue
        col_idx = last_date_col + 1
        ws.cell(row=2, column=col_idx, value=new_month)
        ws.cell(row=3, column=col_idx, value=indicator_header[last_date_col - 1])
        print(f"  → {sheet_name}: {new_month} 컬럼 새로 추가!")

    # KOSIS 지표 컬럼 인덱스 찾기
    kosis_col_idx = None
    for i, val in enumerate(indicator_row):
        if str(val) == kosis_col_name:
            kosis_col_idx = i
            break

    if kosis_col_idx is None:
        print(f"  ⚠ {sheet_name}: KOSIS에서 '{kosis_col_name}' 컬럼을 찾을 수 없어!")
        skipped_sheets.append(sheet_name)
        continue

    # macro 시트 산업분류 빈칸 채우기
    all_rows = list(ws.iter_rows(min_row=4, values_only=True))
    macro_industries = []
    last_industry = ""
    for ws_row in all_rows:
        industry = str(ws_row[0]).strip() if ws_row[0] else ""
        if industry:
            last_industry = industry
        else:
            industry = last_industry
        macro_industries.append(industry)

    # 행별 데이터 매칭
    matched = 0
    for i, ws_row in enumerate(all_rows):
        macro_industry = macro_industries[i]
        macro_size = str(ws_row[1]).strip() if ws_row[1] else ""
        row_idx = i + 4

        for _, kosis_row in kosis_df.iloc[2:].iterrows():
            k_industry = str(kosis_row[0]).strip()
            k_size = str(kosis_row[1]).strip()
            if macro_industry == k_industry and macro_size == k_size:
                value = kosis_row[kosis_col_idx]
                ws.cell(row=row_idx, column=col_idx, value=value)
                matched += 1
                break

    print(f"  ✅ {sheet_name}: {matched}개 행 업데이트 완료")
    updated_sheets.append(sheet_name)

# =============================================
# 5. 원본 파일에 저장
# =============================================
wb.save(MACRO_FILE)

print(f"\n{'=' * 50}")
print(f"완료! {new_month} 데이터가 추가됐어 🎉")
print(f"업데이트된 시트: {len(updated_sheets)}개")
if skipped_sheets:
    print(f"건너뛴 시트: {skipped_sheets}")
print("=" * 50)
