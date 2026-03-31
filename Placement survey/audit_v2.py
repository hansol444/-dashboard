import openpyxl
import re
from collections import defaultdict

TARGET_SHEETS = [
    '1. 지원 RMS', '1-1. 채널 Cut', '1-2. Cubicle Cut', '1-3. Seg Cut',
    '1-4. Seg 시사점', '1-5. 지원 채널 선택 이유',
    '2. 인지 RMS', '2-1. 채널 Cut',
    '3. 재지원 RMS', '3-1. 채널 Cut', '3-2. 재지원 채널 선택 이유',
    'Cover'
]

def get_col_letter(col_num):
    result = ""
    while col_num > 0:
        col_num, remainder = divmod(col_num - 1, 26)
        result = chr(65 + remainder) + result
    return result

def cell_ref(row, col):
    return f"{get_col_letter(col)}{row}"

def parse_countifs_args(inner):
    """Parse COUNTIFS arguments respecting nested parens."""
    parts = []
    depth = 0
    current = ""
    for ch in inner:
        if ch == '(':
            depth += 1
            current += ch
        elif ch == ')':
            depth -= 1
            current += ch
        elif ch == ',' and depth == 0:
            parts.append(current.strip())
            current = ""
        else:
            current += ch
    if current.strip():
        parts.append(current.strip())
    return parts

def extract_col_from_range(rng):
    """Extract column letter from a range like $C$2:$C$442"""
    m = re.findall(r'\$?([A-Z]+)\$?\d+:\$?([A-Z]+)\$?\d+', rng, re.IGNORECASE)
    if m:
        return m[0][0].upper()
    return None

def extract_row_range(rng):
    """Extract start and end rows from range"""
    m = re.findall(r'\$?[A-Z]+\$?(\d+):\$?[A-Z]+\$?(\d+)', rng, re.IGNORECASE)
    if m:
        return (int(m[0][0]), int(m[0][1]))
    return None

def audit_file(filepath, label):
    print(f"\n{'='*80}")
    print(f"AUDITING: {label}")
    print(f"{'='*80}")

    wb = openpyxl.load_workbook(filepath, data_only=False)

    bugs = []
    formula_count = 0
    sheet_formula_counts = {}

    available_sheets = [s for s in wb.sheetnames if s in TARGET_SHEETS]
    print(f"Sheets to audit: {available_sheets}")
    print(f"All sheets in workbook: {wb.sheetnames}")

    for ws_name in available_sheets:
        ws = wb[ws_name]
        sheet_formulas = 0
        formulas_by_type = defaultdict(int)

        for row_cells in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row_cells:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    formula = cell.value
                    formula_count += 1
                    sheet_formulas += 1
                    r = cell.row
                    c = cell.column
                    loc = cell_ref(r, c)

                    # Categorize
                    f_upper = formula.upper()
                    if 'COUNTIFS(' in f_upper:
                        formulas_by_type['COUNTIFS'] += 1
                    elif 'SUMPRODUCT(' in f_upper:
                        formulas_by_type['SUMPRODUCT'] += 1
                    elif 'VLOOKUP(' in f_upper:
                        formulas_by_type['VLOOKUP'] += 1
                    elif '/' in formula:
                        formulas_by_type['Division'] += 1
                    else:
                        formulas_by_type['Other'] += 1

                    # === CHECK 1: Self-reference (circular) ===
                    cell_addr = f"{get_col_letter(c)}{r}"
                    # More careful check - look for the exact cell ref not as part of range
                    # E.g. cell D5 referencing D5 (not D50, not AD5)
                    col_letter = get_col_letter(c)
                    # Pattern: not preceded by letter, exact col+row, not followed by digit
                    pattern_plain = re.compile(
                        r'(?<![A-Z])' + re.escape(col_letter) + str(r) + r'(?!\d)',
                        re.IGNORECASE
                    )
                    pattern_abs = re.compile(
                        r'\$' + re.escape(col_letter) + r'\$' + str(r) + r'(?!\d)',
                        re.IGNORECASE
                    )
                    if pattern_plain.search(formula[1:]) or pattern_abs.search(formula[1:]):
                        # Double check it's not just a range endpoint
                        bugs.append({
                            'sheet': ws_name, 'cell': loc, 'formula': formula,
                            'issue': 'POTENTIAL CIRCULAR REFERENCE - cell may reference itself',
                            'impact': 'Circular reference error or incorrect calculation',
                            'severity': 'HIGH'
                        })

                    # === CHECK 2: COUNTIFS column repetition ===
                    if 'COUNTIFS(' in f_upper:
                        # Find all COUNTIFS calls (handle nested)
                        # Use a more robust extraction
                        idx = 0
                        while True:
                            pos = f_upper.find('COUNTIFS(', idx)
                            if pos == -1:
                                break
                            # Extract the balanced content
                            start = pos + len('COUNTIFS(')
                            depth = 1
                            end = start
                            while end < len(formula) and depth > 0:
                                if formula[end] == '(':
                                    depth += 1
                                elif formula[end] == ')':
                                    depth -= 1
                                end += 1
                            inner = formula[start:end-1]
                            idx = end

                            parts = parse_countifs_args(inner)
                            if len(parts) % 2 != 0:
                                bugs.append({
                                    'sheet': ws_name, 'cell': loc, 'formula': formula,
                                    'issue': f'COUNTIFS has odd number of arguments ({len(parts)})',
                                    'impact': 'Formula error',
                                    'severity': 'HIGH'
                                })
                                continue

                            # Extract range arguments (even indices)
                            range_args = [parts[i] for i in range(0, len(parts), 2)]
                            range_cols = []
                            range_rows = []
                            for rng in range_args:
                                col_l = extract_col_from_range(rng)
                                row_r = extract_row_range(rng)
                                if col_l:
                                    range_cols.append(col_l)
                                if row_r:
                                    range_rows.append(row_r)

                            # Bug: all ranges use same column (when there are 2+ criteria)
                            if len(range_cols) >= 2:
                                unique_cols = set(range_cols)
                                if len(unique_cols) == 1 and len(range_cols) >= 3:
                                    bugs.append({
                                        'sheet': ws_name, 'cell': loc, 'formula': formula,
                                        'issue': f'COUNTIFS: ALL {len(range_cols)} ranges reference column {range_cols[0]} - criteria columns not differentiated',
                                        'impact': 'Wrong count - applying multiple criteria to same column instead of different columns',
                                        'severity': 'CRITICAL'
                                    })

                            # Bug: inconsistent row ranges
                            if len(range_rows) >= 2:
                                unique_ranges = set(range_rows)
                                if len(unique_ranges) > 1:
                                    bugs.append({
                                        'sheet': ws_name, 'cell': loc, 'formula': formula,
                                        'issue': f'COUNTIFS: Inconsistent row ranges across criteria: {unique_ranges}',
                                        'impact': 'Criteria applied to different sized ranges - may cause errors or wrong counts',
                                        'severity': 'MEDIUM'
                                    })

                    # === CHECK 3: SUMPRODUCT array size mismatch ===
                    if 'SUMPRODUCT(' in f_upper:
                        idx = 0
                        while True:
                            pos = f_upper.find('SUMPRODUCT(', idx)
                            if pos == -1:
                                break
                            start = pos + len('SUMPRODUCT(')
                            depth = 1
                            end = start
                            while end < len(formula) and depth > 0:
                                if formula[end] == '(':
                                    depth += 1
                                elif formula[end] == ')':
                                    depth -= 1
                                end += 1
                            inner = formula[start:end-1]
                            idx = end

                            row_ranges = re.findall(r'\$?[A-Z]+\$?(\d+):\$?[A-Z]+\$?(\d+)', inner)
                            if len(row_ranges) >= 2:
                                sizes = set()
                                for s, e in row_ranges:
                                    sizes.add((int(s), int(e)))
                                if len(sizes) > 1:
                                    bugs.append({
                                        'sheet': ws_name, 'cell': loc, 'formula': formula,
                                        'issue': f'SUMPRODUCT: Array size mismatch - ranges: {sizes}',
                                        'impact': '#VALUE! error or incorrect calculation',
                                        'severity': 'HIGH'
                                    })

                    # === CHECK 4: #REF! ===
                    if '#REF!' in formula:
                        bugs.append({
                            'sheet': ws_name, 'cell': loc, 'formula': formula,
                            'issue': 'Broken reference #REF!',
                            'impact': 'Formula produces #REF! error',
                            'severity': 'CRITICAL'
                        })

                    # === CHECK 5: Mismatched parentheses ===
                    open_p = formula.count('(')
                    close_p = formula.count(')')
                    if open_p != close_p:
                        bugs.append({
                            'sheet': ws_name, 'cell': loc, 'formula': formula,
                            'issue': f'Mismatched parentheses: {open_p} open vs {close_p} close',
                            'impact': 'Formula syntax error',
                            'severity': 'HIGH'
                        })

                    # === CHECK 6: Cross-sheet reference to non-existent sheet ===
                    sheet_refs = re.findall(r"'([^']+)'!", formula)
                    for sr in sheet_refs:
                        if sr not in wb.sheetnames and '#REF' not in sr:
                            bugs.append({
                                'sheet': ws_name, 'cell': loc, 'formula': formula,
                                'issue': f'References non-existent sheet: "{sr}"',
                                'impact': '#REF! error',
                                'severity': 'CRITICAL'
                            })

                    # === CHECK 7: Division formulas - check denominators ===
                    if '/' in formula and 'COUNTIFS' not in f_upper and 'SUMPRODUCT' not in f_upper:
                        # Simple division like =A1/B1 or =SOMETHING/SOMETHING
                        # Check if denominator is a cell that could be zero without IFERROR
                        if 'IFERROR' not in f_upper and 'IF(' not in f_upper:
                            # Not protected against division by zero
                            pass  # This is too noisy to flag

                    # === CHECK 8: RMS calculation patterns ===
                    # Look for typical RMS adjusted/unadjusted patterns
                    # These often involve specific calculation logic
                    if 'RMS' in ws_name:
                        # Check for formulas that should have consistent structure
                        pass  # Will analyze patterns separately

        sheet_formula_counts[ws_name] = (sheet_formulas, dict(formulas_by_type))
        formulas_by_type = defaultdict(int)

    # Print summary
    print(f"\nTotal formulas in target sheets: {formula_count}")
    for sn, (cnt, types) in sheet_formula_counts.items():
        print(f"\n  {sn}: {cnt} formulas")
        for t, n in sorted(types.items(), key=lambda x: -x[1]):
            print(f"    {t}: {n}")

    # Deduplicate similar bugs (e.g., same pattern repeated across rows)
    unique_bugs = []
    seen_patterns = set()
    for bug in bugs:
        # Create a pattern key (sheet + issue type + formula pattern)
        # Normalize formula by removing row numbers
        normalized = re.sub(r'\d+', 'N', bug['formula'][:100])
        key = (bug['sheet'], bug['issue'][:50], normalized)
        if key not in seen_patterns:
            seen_patterns.add(key)
            unique_bugs.append(bug)
        else:
            # Find existing and add count
            for ub in unique_bugs:
                norm2 = re.sub(r'\d+', 'N', ub['formula'][:100])
                key2 = (ub['sheet'], ub['issue'][:50], norm2)
                if key2 == key:
                    if 'count' not in ub:
                        ub['count'] = 2
                    else:
                        ub['count'] += 1
                    if 'additional_cells' not in ub:
                        ub['additional_cells'] = [bug['cell']]
                    else:
                        ub['additional_cells'].append(bug['cell'])
                    break

    print(f"\n{'='*80}")
    print(f"UNIQUE BUG PATTERNS FOUND: {len(unique_bugs)} (from {len(bugs)} total instances)")
    print(f"{'='*80}")

    for i, bug in enumerate(unique_bugs, 1):
        f_display = bug['formula'][:400]
        if len(bug['formula']) > 400:
            f_display += '...'
        count = bug.get('count', 1)
        print(f"\n--- Bug Pattern #{i} [{bug.get('severity','?')}] ---")
        print(f"  Sheet: {bug['sheet']}")
        print(f"  Cell:  {bug['cell']}" + (f" (and {count-1} similar cells)" if count > 1 else ""))
        if 'additional_cells' in bug and len(bug['additional_cells']) <= 20:
            print(f"  Also in: {', '.join(bug['additional_cells'][:20])}")
        print(f"  Formula: {f_display}")
        print(f"  Issue: {bug['issue']}")
        print(f"  Impact: {bug['impact']}")

    return bugs, unique_bugs


print("="*80)
print("COMPREHENSIVE FORMULA AUDIT")
print("="*80)

jk_all, jk_unique = audit_file(
    r"c:\Users\ugin35\Desktop\Placement survey 자동화 revive\JK 전체\25Q4_JK 잡플레이스먼트 분석_v0.xlsx",
    "JK File (25Q4_JK)"
)

am_all, am_unique = audit_file(
    r"c:\Users\ugin35\Desktop\Placement survey 자동화 revive\AM 전체\25Q4_AM 잡플레이스먼트 분석_v0.xlsx",
    "AM File (25Q4_AM)"
)

print(f"\n\n{'#'*80}")
print(f"GRAND SUMMARY")
print(f"{'#'*80}")
print(f"JK File: {len(jk_all)} total bug instances, {len(jk_unique)} unique patterns")
print(f"AM File: {len(am_all)} total bug instances, {len(am_unique)} unique patterns")
print(f"TOTAL: {len(jk_all)+len(am_all)} bug instances")
