# -*- coding: utf-8 -*-
"""
JK Placement Survey Automation - R_통합 생성
Usage:
  python run_jk.py                          # 기존 데이터로 R_통합 생성 + 검증
  python run_jk.py --raw RAW.xlsx --quarter 26Q1  # 신규 분기 추가
  python run_jk.py --verify-parse           # 25Q4 Raw 파싱 검증
"""

import sys
import re
import pandas as pd
import openpyxl
from pathlib import Path

# ============================================================
# Configuration
# ============================================================
BASE_DIR = Path(r"C:\Users\ugin35\Desktop\Placement survey 자동화 revive")
JK_DIR = BASE_DIR / "JK 전체"
CLASSIFICATION_FILE = JK_DIR / "25Q4_JK_분류표.xlsx"    # ← 사람이 편집하는 파일

# 25Q3부터 연간 연봉 포맷 사용 (이전은 월소득)
ANNUAL_INCOME_QUARTERS = {'25Q3', '25Q4', '26Q1', '26Q2', '26Q3', '26Q4',
                          '27Q1', '27Q2', '27Q3', '27Q4'}


# ============================================================
# Step 1: Load Classification Tables from 분류표.xlsx
# ============================================================
def load_classifications(filepath):
    """분류표.xlsx에서 6개 시트를 읽어 매핑 딕셔너리로 반환"""
    wb = openpyxl.load_workbook(filepath, data_only=True)

    # --- 1. Channel 분류표: A→E (채널명 → 채널 재분류) ---
    channel_map = {}
    ws = wb['Channel 분류표']
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=5):
        key, val = row[0].value, row[4].value  # A, E
        if key and val and str(val).strip():
            channel_map[str(key).strip()] = str(val).strip()

    # --- 2. 산업 분류표: A→B (산업 응답값 → 9개 산업 Segment) ---
    industry_cubicle = {}
    ws = wb['산업 분류표']
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=2):
        key, val = row[0].value, row[1].value
        if key and val:
            industry_cubicle[str(key).strip()] = str(val).strip()

    # --- 3. 지역 분류표: A→B (일반), D→E (공공) ---
    region_general = {}
    region_public = {}
    ws = wb['지역 분류표']
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=5):
        gk, gv = row[0].value, row[1].value
        pk, pv = row[3].value, row[4].value
        if gk and gv:
            region_general[str(gk).strip()] = str(gv).strip()
        if pk and pv:
            region_public[str(pk).strip()] = str(pv).strip()

    # --- 4. 이유 분류표: A→B ---
    reason_map = {}
    ws = wb['이유 분류표']
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=2):
        key, val = row[0].value, row[1].value
        if key and val:
            reason_map[str(key).strip()] = str(val).strip()

    # --- 5. 산업직무소득 Seg ---
    seg_industry = {}
    seg_job = {}
    seg_income_annual = {}
    seg_income_monthly = {}
    seg_region = {}

    ws = wb['산업직무소득 Seg']
    for row in ws.iter_rows(min_row=3, min_col=1, max_col=12):
        # A→B: 산업 seg
        if row[0].value and row[1].value:
            seg_industry[str(row[0].value).strip()] = str(row[1].value).strip()
        # D→E: 직무 seg
        if row[3].value and row[4].value:
            seg_job[str(row[3].value).strip()] = str(row[4].value).strip()
        # G→I: 연간 연봉 → 소득수준
        if row[6].value and row[8].value:
            seg_income_annual[str(row[6].value).strip()] = str(row[8].value).strip()
        # H→I: 월간 소득 → 소득수준
        if row[7].value and row[8].value:
            seg_income_monthly[str(row[7].value).strip()] = str(row[8].value).strip()
        # K→L: 지역 seg
        if row[10].value and row[11].value:
            seg_region[str(row[10].value).strip()] = str(row[11].value).strip()

    # --- 6. Cubicle 규칙: A=산업Group, B=연령라벨, C=포함연령값, D=지역매핑 ---
    cubicle_rules = {}   # {산업Group: {frozenset(연령값): (연령라벨, 지역매핑)}}
    ws = wb['Cubicle 규칙']
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=4):
        ind_grp = row[0].value
        age_label = row[1].value
        age_values = row[2].value
        region_type = row[3].value
        if not ind_grp or not age_label or not age_values or not region_type:
            continue
        ind_grp = str(ind_grp).strip()
        age_label = str(age_label).strip()
        region_type = str(region_type).strip()
        ages = frozenset(a.strip() for a in str(age_values).split(','))

        if ind_grp not in cubicle_rules:
            cubicle_rules[ind_grp] = {}
        cubicle_rules[ind_grp][ages] = (age_label, region_type)

    wb.close()
    return {
        'channel': channel_map,
        'industry_cubicle': industry_cubicle,
        'region_general': region_general,
        'region_public': region_public,
        'reason': reason_map,
        'seg_industry': seg_industry,
        'seg_job': seg_job,
        'seg_income_annual': seg_income_annual,
        'seg_income_monthly': seg_income_monthly,
        'seg_region': seg_region,
        'cubicle_rules': cubicle_rules,
    }


# ============================================================
# Step 2: Read Base Data (이전 R_통합 결과)
# ============================================================
# R_통합 Excel 헤더 → 내부 컬럼명 매핑
BASE_HEADER_MAP = {
    'no.': 'no', 'Quarter': 'quarter',
    'SQ1 성별': 'gender', 'SQ2_2 연령': 'age', 'SQ3 거주지': 'region',
    'SQ4 직업': 'occupation', 'SQ7 산업': 'industry', 'SQ8 직무': 'job_function',
    'SQ9 기업 규모': 'company_size', 'SQ10 기업유형': 'company_type',
    'SQ11 연봉': 'income',
    'Why 지원 1': 'why_apply_1', 'Why 지원 2': 'why_apply_2', 'Why 지원 3': 'why_apply_3',
    'Why 재지원 1': 'why_reuse_1', 'Why 재지원 2': 'why_reuse_2', 'Why 재지원 3': 'why_reuse_3',
    '인지 채널': 'channel_aware', '지원 채널': 'channel_apply', '재사용 채널': 'channel_reuse',
}

# Raw 컬럼만 (분류 컬럼은 재계산하므로 읽지 않음)
RAW_FIELD_NAMES = list(BASE_HEADER_MAP.values())


def read_base_data(filepath):
    """이전 R_통합 결과에서 Raw 컬럼만 읽기 (분류 컬럼은 재계산)"""
    df = pd.read_excel(filepath, sheet_name='R_통합', engine='openpyxl')

    # 헤더를 내부 컬럼명으로 변환
    rename = {k: v for k, v in BASE_HEADER_MAP.items() if k in df.columns}
    df = df.rename(columns=rename)

    # Raw 컬럼만 남기기
    keep = [c for c in RAW_FIELD_NAMES if c in df.columns]
    df = df[keep].copy()

    # NaN → 빈 문자열
    df = df.fillna('')
    for col in df.columns:
        df[col] = df[col].astype(str).str.strip()
        df[col] = df[col].replace('nan', '')

    print(f"  Base data: {len(df)} rows from {filepath.name}")
    return df


# ============================================================
# Step 3: Parse New Quarter Raw File
# ============================================================
def strip_prefix(text):
    """설문 응답 접두사 제거: '  2) 사람인' → '사람인'"""
    if not text:
        return ''
    text = str(text)
    m = re.match(r'^\s*\d+\)\s*', text)
    if m:
        return text[m.end():]
    return text.strip()


# 25Q3+ Raw 파일 컬럼 매핑
RAW_FILE_COLUMNS = [
    (2, 'gender', True),
    (5, 'age', True),
    (6, 'region', True),
    (7, 'occupation', True),
    (25, 'industry', True),
    (27, 'job_function', True),
    (29, 'company_size', False),   # 접두사 유지
    (30, 'company_type', True),
    (31, 'income', True),
    (36, 'channel_aware', True),
    (39, 'channel_apply', True),
    (41, 'why_apply_1', True),
    (42, 'why_apply_2', True),
    (43, 'why_apply_3', True),
    (44, 'channel_reuse', True),
    (45, 'why_reuse_1', True),
    (46, 'why_reuse_2', True),
    (47, 'why_reuse_3', True),
]


def parse_raw_file(filepath, quarter):
    """신규 분기 Raw 파일 파싱 (25Q3+ 포맷)"""
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    ws = wb['String']

    data = []
    for row in ws.iter_rows(min_row=2, max_col=52):
        if not row[0].value:
            break
        row_data = {'no': '', 'quarter': quarter}
        for col_num, field, do_strip in RAW_FILE_COLUMNS:
            val = row[col_num - 1].value
            if val:
                row_data[field] = strip_prefix(str(val)) if do_strip else str(val).strip()
            else:
                row_data[field] = ''
        data.append(row_data)

    wb.close()
    print(f"  New quarter raw data ({quarter}): {len(data)} rows parsed")
    return pd.DataFrame(data)


# ============================================================
# Step 4: Classification Logic (all driven by 분류표.xlsx)
# ============================================================
def classify_channel(raw_channel, channel_map):
    if not raw_channel:
        return '미분류'
    return channel_map.get(raw_channel, '미분류')


def classify_industry_group(industry, industry_cubicle):
    if not industry:
        return '기타산업'
    return industry_cubicle.get(industry, '기타산업')


def classify_cubicle(industry_group, age, region, cubicle_rules, region_general, region_public):
    """Cubicle 규칙 시트 기반으로 산업>연령>지역 Group 계산

    Returns: (age_group, region_group, cubicle)
    """
    if industry_group == '기타산업' or not industry_group:
        return '기타', '전국', f'{industry_group} 기타 전국'

    rules = cubicle_rules.get(industry_group)
    if not rules:
        return '기타', '전국', f'{industry_group} 기타 전국'

    # 연령값으로 매칭되는 규칙 찾기
    age_label = ''
    region_type = '전국'
    for age_set, (label, rtype) in rules.items():
        if age in age_set:
            age_label = label
            region_type = rtype
            break

    if not age_label:
        return '기타', '전국', f'{industry_group} 기타 전국'

    # 산업>연령 Group 생성: (스)2030 형태 (기타산업은 "기타")
    prefix = industry_group[0]
    age_group = f'({prefix}){age_label}'

    # 지역 매핑 적용
    if region_type == 'general':
        region_group = region_general.get(region, '미분류')
    elif region_type == 'public':
        region_group = region_public.get(region, '미분류')
    else:
        region_group = '전국'

    cubicle = f'{industry_group} {age_group} {region_group}'
    return age_group, region_group, cubicle


def classify_reason(reason_text, reason_map):
    if not reason_text:
        return ''
    return reason_map.get(reason_text, '')


def classify_income(income, quarter, seg_annual, seg_monthly):
    if not income:
        return ''
    if quarter in ANNUAL_INCOME_QUARTERS:
        cleansed = income.replace(' ~ ', '-')
        return seg_annual.get(cleansed, '')
    else:
        return seg_monthly.get(income, '')


def apply_all_classifications(df, cls):
    """모든 분류 로직을 DataFrame에 적용"""
    ch = cls['channel']
    ic = cls['industry_cubicle']
    rg = cls['region_general']
    rp = cls['region_public']
    rm = cls['reason']
    si = cls['seg_industry']
    sj = cls['seg_job']
    sa = cls['seg_income_annual']
    sm = cls['seg_income_monthly']
    sr = cls['seg_region']
    cr = cls['cubicle_rules']

    # --- 연봉 Cleansing ---
    df['income_cleansed'] = df.apply(
        lambda r: r['income'].replace(' ~ ', '-')
        if r['quarter'] in ANNUAL_INCOME_QUARTERS and r['income']
        else '', axis=1
    )

    # --- 이유 분류 ---
    for suffix in ['1', '2', '3']:
        df[f'why_apply_cat_{suffix}'] = df[f'why_apply_{suffix}'].map(
            lambda x: classify_reason(x, rm))
        df[f'why_reuse_cat_{suffix}'] = df[f'why_reuse_{suffix}'].map(
            lambda x: classify_reason(x, rm))

    # --- 채널 재분류 ---
    df['channel_aware_rms'] = df['channel_aware'].map(lambda x: classify_channel(x, ch))
    df['channel_apply_rms'] = df['channel_apply'].map(lambda x: classify_channel(x, ch))
    df['channel_reuse_rms'] = df['channel_reuse'].map(lambda x: classify_channel(x, ch))

    # --- Cubicle (분류표.xlsx Cubicle 규칙 시트 기반) ---
    df['industry_group'] = df['industry'].map(lambda x: classify_industry_group(x, ic))

    cubicle_results = df.apply(
        lambda r: classify_cubicle(
            r['industry_group'], r['age'], r['region'], cr, rg, rp
        ), axis=1, result_type='expand'
    )
    df['age_group'] = cubicle_results[0]
    df['region_group'] = cubicle_results[1]
    df['cubicle'] = cubicle_results[2]

    # --- Seg 라벨 ---
    df['seg_industry'] = df.apply(
        lambda r: '기타' if not r['channel_apply']
        else si.get(r['industry'], '미분류') if r['industry'] else '미분류',
        axis=1
    )
    df['seg_job'] = df['job_function'].map(
        lambda x: sj.get(x, '기타') if x else '기타')
    df['seg_income'] = df.apply(
        lambda r: classify_income(r['income'], r['quarter'], sa, sm), axis=1)
    df['seg_region'] = df['region'].map(
        lambda x: sr.get(x, '') if x else '')

    return df


# ============================================================
# Step 5: Output
# ============================================================
OUTPUT_COLUMNS = [
    ('no', 'no.'),
    ('quarter', 'Quarter'),
    ('gender', 'SQ1 성별'),
    ('age', 'SQ2_2 연령'),
    ('region', 'SQ3 거주지'),
    ('occupation', 'SQ4 직업'),
    ('industry', 'SQ7 산업'),
    ('job_function', 'SQ8 직무'),
    ('company_size', 'SQ9 기업 규모'),
    ('company_type', 'SQ10 기업유형'),
    ('income', 'SQ11 연봉'),
    ('income_cleansed', '연봉 Cleansing'),
    ('why_apply_1', 'Why 지원 1'),
    ('why_apply_2', 'Why 지원 2'),
    ('why_apply_3', 'Why 지원 3'),
    ('why_apply_cat_1', 'Why 지원 분류 1'),
    ('why_apply_cat_2', 'Why 지원 분류 2'),
    ('why_apply_cat_3', 'Why 지원 분류 3'),
    ('why_reuse_1', 'Why 재지원 1'),
    ('why_reuse_2', 'Why 재지원 2'),
    ('why_reuse_3', 'Why 재지원 3'),
    ('why_reuse_cat_1', 'Why 재지원 분류 1'),
    ('why_reuse_cat_2', 'Why 재지원 분류 2'),
    ('why_reuse_cat_3', 'Why 재지원 분류 3'),
    ('channel_aware', '인지 채널'),
    ('channel_aware_rms', '인지 RMS 재분류'),
    ('channel_apply', '지원 채널'),
    ('channel_apply_rms', '지원 RMS 재분류'),
    ('channel_reuse', '재사용 채널'),
    ('channel_reuse_rms', '재사용 RMS 재분류'),
    ('industry_group', '산업 Group'),
    ('age_group', '산업>연령 Group'),
    ('region_group', '산업>연령>지역 Group'),
    ('cubicle', '최종 Cubicle'),
    ('seg_industry', 'Seg 산업'),
    ('seg_job', 'Seg 직무'),
    ('seg_income', 'Seg 소득수준'),
    ('seg_region', 'Seg 지역'),
]


def write_output(df, filepath):
    col_keys = [c[0] for c in OUTPUT_COLUMNS]
    col_headers = [c[1] for c in OUTPUT_COLUMNS]

    out_df = df[col_keys].copy()
    out_df.columns = col_headers

    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        out_df.to_excel(writer, sheet_name='R_통합', index=False)

    print(f"  Output: {filepath}")
    print(f"  Rows: {len(out_df)}, Columns: {len(out_df.columns)}")


# ============================================================
# Verification
# ============================================================
def verify_against_existing(df, filepath):
    """기존 Excel R_22Q2-25Q4와 Python 결과 대조"""
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    ws = wb['R_22Q2-25Q4']

    check_cols = {
        26: ('channel_aware_rms', 'RMS 인지 채널'),
        28: ('channel_apply_rms', 'RMS 지원 채널'),
        30: ('channel_reuse_rms', 'RMS 재사용 채널'),
        31: ('industry_group', '산업 Group'),
        34: ('cubicle', '최종 Cubicle'),
        35: ('seg_industry', 'Seg 산업'),
        36: ('seg_job', 'Seg 직무'),
        37: ('seg_income', 'Seg 소득수준'),
        38: ('seg_region', 'Seg 지역'),
    }

    mismatches = {name: 0 for _, (_, name) in check_cols.items()}
    total = 0
    sample_errors = {name: [] for _, (_, name) in check_cols.items()}

    for r_idx, row in enumerate(ws.iter_rows(min_row=3, max_col=38)):
        if r_idx >= len(df):
            break
        if not row[1].value:
            break
        total += 1
        quarter = row[1].value

        for col_idx, (py_col, label) in check_cols.items():
            excel_val = str(row[col_idx - 1].value).strip() if row[col_idx - 1].value else ''
            py_val = str(df.iloc[r_idx][py_col]).strip() if df.iloc[r_idx][py_col] else ''
            if excel_val != py_val:
                mismatches[label] += 1
                if len(sample_errors[label]) < 3:
                    sample_errors[label].append(
                        f"  Row {r_idx+3}: Excel='{excel_val}' vs Python='{py_val}'"
                        f" (quarter={quarter})"
                    )

    wb.close()

    print(f"\n=== Verification: {total} rows ===")
    all_ok = True
    for label, count in mismatches.items():
        if count > 0:
            print(f"  {label}: {count} mismatches ({count/total*100:.1f}%)")
            for err in sample_errors[label]:
                print(err)
            all_ok = False
        else:
            print(f"  {label}: OK")

    if all_ok:
        print("\n  All match!")
    return all_ok


def verify_parsing(parsed_df, historical_df, quarter):
    """Raw 파싱 결과와 기존 R_통합 대조"""
    existing = historical_df[historical_df['quarter'] == quarter].reset_index(drop=True)
    parsed = parsed_df.reset_index(drop=True)

    if len(existing) == 0:
        print(f"  기존 데이터에 {quarter} 없음 — 신규 분기")
        return True

    print(f"  기존 {len(existing)}행 vs 파싱 {len(parsed)}행")
    check_fields = [
        'gender', 'age', 'region', 'occupation', 'industry',
        'job_function', 'income', 'channel_aware', 'channel_apply',
        'channel_reuse', 'why_apply_1', 'why_apply_2', 'why_apply_3',
        'why_reuse_1', 'why_reuse_2', 'why_reuse_3',
    ]

    n = min(len(existing), len(parsed))
    for field in check_fields:
        mismatches = 0
        samples = []
        for i in range(n):
            ev = str(existing.iloc[i][field]).strip() if existing.iloc[i][field] else ''
            pv = str(parsed.iloc[i][field]).strip() if parsed.iloc[i][field] else ''
            if ev != pv:
                mismatches += 1
                if len(samples) < 2:
                    samples.append(f"    Row {i}: existing='{ev}' vs parsed='{pv}'")
        if mismatches > 0:
            print(f"  {field}: {mismatches} mismatches ({mismatches/n*100:.1f}%)")
            for s in samples:
                print(s)
        else:
            print(f"  {field}: OK")
    return True


# ============================================================
# Main
# ============================================================
def main():
    import argparse
    parser = argparse.ArgumentParser(description='JK Placement Survey - R_통합 생성')
    parser.add_argument('--base', type=str, required=True,
                        help='이전 R_통합 결과 파일 (과거 데이터 원본)')
    parser.add_argument('--raw', type=str,
                        help='신규 분기 Raw 파일 경로')
    parser.add_argument('--quarter', type=str,
                        help='신규 분기 이름 (예: 26Q1)')
    parser.add_argument('--output', type=str,
                        help='출력 파일 경로 (기본: JK 전체/{quarter}_JK_결과.xlsx)')
    args = parser.parse_args()

    print("=" * 60)
    print("JK Placement Survey - R_통합 생성")
    print("=" * 60)

    # Step 1: 분류표 로드
    print(f"\n[Step 1] Loading: {CLASSIFICATION_FILE.name}")
    cls = load_classifications(CLASSIFICATION_FILE)
    print(f"  Channel: {len(cls['channel'])} | Industry: {len(cls['industry_cubicle'])}")
    print(f"  Reason: {len(cls['reason'])} | Cubicle rules: {sum(len(v) for v in cls['cubicle_rules'].values())}")

    # Step 2: 이전 R_통합에서 과거 데이터 읽기
    base_path = Path(args.base)
    if not base_path.exists():
        print(f"  ERROR: {base_path} not found")
        return
    print(f"\n[Step 2] Reading base: {base_path.name}")
    df = read_base_data(base_path)
    print(f"  Quarters: {sorted(df['quarter'].unique().tolist())}")

    # Step 3: 신규 Raw 파싱
    if args.raw and args.quarter:
        raw_path = Path(args.raw)
        if not raw_path.exists():
            print(f"  ERROR: {raw_path} not found")
            return
        print(f"\n[Step 3] Parsing {args.quarter}...")
        new_df = parse_raw_file(raw_path, args.quarter)
        df = df[df['quarter'] != args.quarter]
        df = pd.concat([df, new_df], ignore_index=True)
        print(f"  Total: {len(df)} rows")

    # Step 4: 분류 적용
    print("\n[Step 4] Classifying...")
    df = apply_all_classifications(df, cls)

    # 미분류 상세 리포트
    # "기타산업", "기타" 등 정상 기본값은 제외하고, 진짜 미등록 값만 잡음
    print("\n" + "=" * 60)
    print("[미분류 Report]")
    print("=" * 60)

    # (분류 결과 컬럼, 원본 값 컬럼, 미분류 판정값, 수정할 분류표)
    # 기타산업/기타 = 정상 기본값이므로 체크 대상 아님
    checks = [
        ('channel_aware_rms', 'channel_aware', '미분류', 'Channel 분류표', '인지 채널'),
        ('channel_apply_rms', 'channel_apply', '미분류', 'Channel 분류표', '지원 채널'),
        ('channel_reuse_rms', 'channel_reuse', '미분류', 'Channel 분류표', '재사용 채널'),
        ('seg_industry', 'industry', '미분류', '산업직무소득 Seg (산업)', 'Seg 산업'),
        ('seg_income', 'income', '', '산업직무소득 Seg (소득)', 'Seg 소득수준'),
    ]

    has_issues = False
    for result_col, raw_col, bad_val, table_name, label in checks:
        mask = (df[result_col] == bad_val) & (df[raw_col] != '')
        problem_df = df[mask]

        if len(problem_df) == 0:
            continue

        has_issues = True
        grouped = problem_df.groupby(raw_col).apply(
            lambda g: g.index.tolist(), include_groups=False)

        print(f"\n  {label}: {len(problem_df)}건 미분류 → {table_name}에 추가 필요")
        for raw_val, indices in grouped.items():
            row_nums = [str(idx + 2) for idx in indices[:5]]
            row_str = ', '.join(row_nums)
            if len(indices) > 5:
                row_str += f' 외 {len(indices)-5}건'
            print(f"    - \"{raw_val}\" ({len(indices)}건, 결과 행: {row_str})")

    if not has_issues:
        print("  미분류 없음!")

    # Step 5: 출력
    if args.output:
        output_path = Path(args.output)
    elif args.quarter:
        output_path = JK_DIR / f"{args.quarter}_JK_결과.xlsx"
    else:
        output_path = JK_DIR / "JK_결과.xlsx"

    print(f"\n[Step 5] Writing output...")
    write_output(df, output_path)

    print("\nDone!")


if __name__ == '__main__':
    main()
