# -*- coding: utf-8 -*-
"""
AM Placement Survey Automation - R_통합 생성
Usage:
  python run_am.py --base "AM 전체/25Q4_AM_결과.xlsx" --raw RAW.xlsx --quarter 26Q1
"""
import sys
import re
import pandas as pd
import openpyxl
from pathlib import Path

# ============================================================
# Configuration
# ============================================================
BASE_DIR = Path(r"C:\Users\olivia408\projects\-dashboard\Placement survey")
AM_DIR = BASE_DIR / "AM 전체"
CLASSIFICATION_FILE = AM_DIR / "25Q4_AM_분류표.xlsx"

ANNUAL_INCOME_QUARTERS = set()  # AM은 전 분기 월소득(DQ3) 사용


# ============================================================
# Step 1: Load Classification Tables from 분류표.xlsx
# ============================================================
def load_classifications(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)

    # Channel 분류표: A(채널명) → D(채널 재분류)
    channel_map = {}
    ws = wb['Channel 분류표']
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=4):
        key, val = row[0].value, row[3].value  # A=raw, D=target
        if key and val and str(val).strip():
            channel_map[str(key).strip()] = str(val).strip()

    # 이유 분류표: A→B
    reason_map = {}
    ws = wb['이유 분류표']
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=2):
        key, val = row[0].value, row[1].value
        if key and val:
            reason_map[str(key).strip()] = str(val).strip()

    # Segment 분류표
    seg_job = {}         # A→B: 직무
    seg_contract = {}    # D→E: 근무형태(계약기간)
    seg_income = {}      # G→H: 소득
    seg_region = {}      # K→L: 지역

    ws = wb['Segment 분류표']
    for row in ws.iter_rows(min_row=3, min_col=1, max_col=12):
        if row[0].value and row[1].value:
            seg_job[str(row[0].value).strip()] = str(row[1].value).strip()
        if row[3].value and row[4].value:
            seg_contract[str(row[3].value).strip()] = str(row[4].value).strip()
        if row[6].value and row[7].value:
            # 쉼표 구분 복합값 처리
            raw_income = str(row[6].value).strip()
            val_income = str(row[7].value).strip()
            if ',' in raw_income:
                for part in raw_income.split(','):
                    part = part.strip()
                    if part:
                        seg_income[part] = val_income
            else:
                seg_income[raw_income] = val_income
        if row[10].value and row[11].value:
            seg_region[str(row[10].value).strip()] = str(row[11].value).strip()

    # Cubicle 연령 규칙: 성별 + 연령대 → 연령Group
    cubicle_age_rules = {}  # {(성별, 연령대): 연령Group}
    ws = wb['Cubicle 연령 규칙']
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=3):
        gender = str(row[0].value).strip() if row[0].value else ''
        ages_str = str(row[1].value).strip() if row[1].value else ''
        age_grp = str(row[2].value).strip() if row[2].value else ''
        if gender and ages_str and age_grp:
            for age in ages_str.split(','):
                cubicle_age_rules[(gender, age.strip())] = age_grp

    # Cubicle 규칙: (성별, 연령Group, 거주지) → 거주지재분류
    cubicle_region_rules = {}
    ws = wb['Cubicle 규칙']
    current_gender = ''
    current_age = ''
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=5):
        g = str(row[0].value).strip() if row[0].value else ''
        a = str(row[1].value).strip() if row[1].value else ''
        r = str(row[2].value).strip() if row[2].value else ''
        r2 = str(row[3].value).strip() if row[3].value else ''
        if g:
            current_gender = g
        if a:
            current_age = a
        if r and current_gender and current_age:
            cubicle_region_rules[(current_gender, current_age, r)] = r2 if r2 else r

    wb.close()
    return {
        'channel': channel_map,
        'reason': reason_map,
        'seg_job': seg_job,
        'seg_contract': seg_contract,
        'seg_income': seg_income,
        'seg_region': seg_region,
        'cubicle_age_rules': cubicle_age_rules,
        'cubicle_region_rules': cubicle_region_rules,
    }


# ============================================================
# Step 2: Read Base Data
# ============================================================
BASE_HEADER_MAP = {
    'no.': 'no', 'Quarter': 'quarter',
    'SQ1 성별': 'gender', 'SQ2 연령대': 'age_band', 'SQ3 거주지': 'region',
    'SQ6 직무': 'job_function', 'SQ7 기간': 'contract_period',
    'DQ3 가계소득': 'income',
    'Why 지원 1': 'why_apply_1', 'Why 지원 2': 'why_apply_2', 'Why 지원 3': 'why_apply_3',
    'Why 재지원 1': 'why_reuse_1', 'Why 재지원 2': 'why_reuse_2', 'Why 재지원 3': 'why_reuse_3',
    '인지 채널': 'channel_aware', '지원 채널': 'channel_apply', '재사용 채널': 'channel_reuse',
}
RAW_FIELDS = list(BASE_HEADER_MAP.values())


def read_base_data(filepath):
    df = pd.read_excel(filepath, sheet_name='R_통합', engine='openpyxl')
    rename = {k: v for k, v in BASE_HEADER_MAP.items() if k in df.columns}
    df = df.rename(columns=rename)
    keep = [c for c in RAW_FIELDS if c in df.columns]
    df = df[keep].copy().fillna('')
    for col in df.columns:
        df[col] = df[col].astype(str).str.strip().replace('nan', '')
    print(f"  Base data: {len(df)} rows from {filepath.name}")
    return df


def read_historical_raw(filepath):
    """기존 분석 Excel R_22.2Q-25.4Q에서 Raw 컬럼 읽기 (초기 1회용)"""
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    ws = wb['R_22.2Q-25.4Q']

    # AM R_통합 컬럼 매핑 (1-based)
    col_map = {
        1: 'no', 2: 'quarter', 3: 'gender', 5: 'age_band', 6: 'region',
        7: 'job_function', 8: 'contract_period', 9: 'income',
        10: 'why_apply_1', 11: 'why_apply_2', 12: 'why_apply_3',
        13: 'why_reuse_1', 14: 'why_reuse_2', 15: 'why_reuse_3',
        22: 'channel_aware', 23: 'channel_apply', 24: 'channel_reuse',
        31: '_orig_cubicle',  # AE열: 기존 cubicle 라벨 (재계산 검증용)
    }

    data = []
    for row in ws.iter_rows(min_row=3, max_col=31):
        q = row[1].value
        if not q:
            break
        row_data = {}
        for ci, name in col_map.items():
            val = row[ci - 1].value
            row_data[name] = str(val).strip() if val else ''
        data.append(row_data)

    wb.close()
    print(f"  Historical raw: {len(data)} rows")
    return pd.DataFrame(data)


# ============================================================
# Step 3: Parse New Quarter Raw File
# ============================================================
def strip_prefix(text):
    if not text:
        return ''
    text = str(text)
    m = re.match(r'^\s*\d+\)\s*', text)
    if m:
        return text[m.end():]
    return text.strip()


# 25Q3+ AM Raw 컬럼 매핑
RAW_FILE_COLUMNS = [
    (2, 'gender', True),           # SQ1
    (5, 'age_band', True),         # SQ2
    (6, 'region', True),           # SQ3
    (8, 'job_function', True),     # SQ6_1 직무
    (9, 'contract_period', True),  # SQ7
    (36, 'channel_aware', True),   # B3 인지
    (38, 'channel_apply', True),   # B5 지원
    (40, 'channel_reuse', True),   # B7 재사용
    (42, 'why_apply_1', True),     # B6_1
    (43, 'why_apply_2', True),     # B6_2
    (44, 'why_apply_3', True),     # B6_3
    (45, 'why_reuse_1', True),     # B8_1
    (46, 'why_reuse_2', True),     # B8_2
    (47, 'why_reuse_3', True),     # B8_3
]


def parse_raw_file(filepath, quarter):
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    ws = wb['String']

    # DQ3 위치 찾기 (헤더에서)
    headers = {}
    for row in ws.iter_rows(min_row=1, max_row=1, max_col=60):
        for cell in row:
            if cell.value:
                headers[str(cell.value).strip()] = cell.column
    dq3_col = headers.get('DQ3', None)

    data = []
    for row in ws.iter_rows(min_row=2, max_col=60):
        if not row[0].value:
            break
        row_data = {'no': '', 'quarter': quarter}
        for col_num, field, do_strip in RAW_FILE_COLUMNS:
            val = row[col_num - 1].value if col_num - 1 < len(row) else None
            if val:
                row_data[field] = strip_prefix(str(val)) if do_strip else str(val).strip()
            else:
                row_data[field] = ''
        # DQ3 income
        if dq3_col and dq3_col - 1 < len(row):
            val = row[dq3_col - 1].value
            row_data['income'] = strip_prefix(str(val)) if val else ''
        else:
            row_data['income'] = ''
        data.append(row_data)

    wb.close()
    print(f"  New quarter ({quarter}): {len(data)} rows parsed")
    return pd.DataFrame(data)


# ============================================================
# Step 4: Classification Logic
# ============================================================
def classify_channel(raw, channel_map):
    if not raw:
        return '미분류'
    return channel_map.get(raw, '미분류')


def classify_reason(text, reason_map):
    if not text:
        return ''
    return reason_map.get(text, '')


def classify_cubicle(gender, age_band, region, age_rules, region_rules):
    """AM Cubicle: 성별 + 연령Group + 거주지재분류"""
    age_grp = age_rules.get((gender, age_band), '')
    if not age_grp:
        return '', '', f'{gender}{age_grp}{region}'

    region2 = region_rules.get((gender, age_grp, region), region)
    label = f'{gender}{age_grp}{region2}'
    return age_grp, region2, label


def classify_income_am(income_str):
    """AM 소득 분류: 문자열에서 금액 추출하여 분류"""
    if not income_str:
        return ''
    s = income_str
    # 숫자 추출
    if '200만원 미만' in s or '100만원' in s:
        return '저소득'
    if '200' in s and '300' in s:
        return '저소득'
    if '300' in s and '400' in s:
        return '중소득'
    if '400' in s and '500' in s:
        return '중소득'
    if '500' in s and '600' in s:
        return '중소득'
    if '600' in s and '700' in s:
        return '고소득'
    if '700' in s and '800' in s:
        return '고소득'
    if '800' in s and '900' in s:
        return '고소득'
    if '900' in s or '1000' in s or '1,000' in s:
        return '초고소득'
    return ''


def apply_all_classifications(df, cls):
    ch = cls['channel']
    rm = cls['reason']
    sj = cls['seg_job']
    sc_map = cls['seg_contract']
    si = cls['seg_income']
    sr = cls['seg_region']
    ar = cls['cubicle_age_rules']
    rr = cls['cubicle_region_rules']

    # 이유 분류
    for sfx in ['1', '2', '3']:
        df[f'why_apply_cat_{sfx}'] = df[f'why_apply_{sfx}'].map(lambda x: classify_reason(x, rm))
        df[f'why_reuse_cat_{sfx}'] = df[f'why_reuse_{sfx}'].map(lambda x: classify_reason(x, rm))

    # 채널 재분류
    df['channel_aware_rms'] = df['channel_aware'].map(lambda x: classify_channel(x, ch))
    df['channel_apply_rms'] = df['channel_apply'].map(lambda x: classify_channel(x, ch))
    df['channel_reuse_rms'] = df['channel_reuse'].map(lambda x: classify_channel(x, ch))

    # 지역 Seg
    df['seg_region'] = df['region'].map(lambda x: sr.get(x, x) if x else '')

    # Cubicle용 지역 (Seg 지역 → Cubicle 5그룹: 서울/경기강원권/영남권/충청권/호남권)
    CUB_REGION = {'서울특별시': '서울', '서울': '서울', '제주': '호남권', '제주도': '호남권'}
    df['_cub_region'] = df['seg_region'].map(lambda x: CUB_REGION.get(x, x))

    # Cubicle: 기존 데이터에 _orig_cubicle이 있으면 그대로 사용, 없으면 재계산
    has_orig = '_orig_cubicle' in df.columns
    cub_results = df.apply(
        lambda r: classify_cubicle(r['gender'], r['age_band'], r['_cub_region'], ar, rr),
        axis=1, result_type='expand'
    )
    df.drop(columns=['_cub_region'], inplace=True)
    df['age_group'] = cub_results[0]
    df['region_group'] = cub_results[1]
    df['cubicle'] = cub_results[2]

    # 기존 cubicle 라벨이 있으면 덮어쓰기 (과거 데이터 정합성 유지)
    if has_orig:
        mask = df['_orig_cubicle'].astype(str).str.strip() != ''
        df.loc[mask, 'cubicle'] = df.loc[mask, '_orig_cubicle'].str.strip()
        df.drop(columns=['_orig_cubicle'], inplace=True)

    # Seg
    df['seg_job'] = df['job_function'].map(lambda x: sj.get(x, '기타') if x else '기타')
    df['seg_contract'] = df['contract_period'].map(lambda x: sc_map.get(x, '기타') if x else '기타')

    # 소득: 분류표 매핑 → 실패시 금액 기반 분류
    def classify_income_am(income):
        if not income:
            return ''
        # 1차: 분류표 exact match
        mapped = si.get(income, '')
        if mapped:
            return mapped
        # 2차: 금액 추출 후 범위 매핑
        nums = re.findall(r'(\d+)', income.replace(',', ''))
        if not nums:
            return ''
        first_num = int(nums[0])
        if first_num < 300:
            return '저소득'
        elif first_num < 600:
            return '중소득'
        elif first_num < 900:
            return '고소득'
        else:
            return '고소득'
    df['seg_income'] = df['income'].map(classify_income_am)

    return df


# ============================================================
# Step 5: Output
# ============================================================
OUTPUT_COLUMNS = [
    ('no', 'no.'), ('quarter', 'Quarter'),
    ('gender', 'SQ1 성별'), ('age_band', 'SQ2 연령대'), ('region', 'SQ3 거주지'),
    ('job_function', 'SQ6 직무'), ('contract_period', 'SQ7 기간'), ('income', 'DQ3 가계소득'),
    ('why_apply_1', 'Why 지원 1'), ('why_apply_2', 'Why 지원 2'), ('why_apply_3', 'Why 지원 3'),
    ('why_apply_cat_1', 'Why 지원 분류 1'), ('why_apply_cat_2', 'Why 지원 분류 2'),
    ('why_apply_cat_3', 'Why 지원 분류 3'),
    ('why_reuse_1', 'Why 재지원 1'), ('why_reuse_2', 'Why 재지원 2'), ('why_reuse_3', 'Why 재지원 3'),
    ('why_reuse_cat_1', 'Why 재지원 분류 1'), ('why_reuse_cat_2', 'Why 재지원 분류 2'),
    ('why_reuse_cat_3', 'Why 재지원 분류 3'),
    ('channel_aware', '인지 채널'), ('channel_aware_rms', '인지 RMS 재분류'),
    ('channel_apply', '지원 채널'), ('channel_apply_rms', '지원 RMS 재분류'),
    ('channel_reuse', '재사용 채널'), ('channel_reuse_rms', '재사용 RMS 재분류'),
    ('age_group', '연령 Group'), ('region_group', '거주지 재분류'),
    ('cubicle', '최종 Cubicle'),
    ('seg_job', 'Seg 직무'), ('seg_contract', 'Seg 근무형태'),
    ('seg_income', 'Seg 소득'), ('seg_region', 'Seg 지역'),
]


def write_output(df, filepath):
    col_keys = [c[0] for c in OUTPUT_COLUMNS]
    col_headers = [c[1] for c in OUTPUT_COLUMNS]
    out_df = df[col_keys].copy()
    out_df.columns = col_headers
    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        out_df.to_excel(writer, sheet_name='R_통합', index=False)
    print(f"  Output: {filepath}")
    print(f"  Rows: {len(out_df)}, Columns: {len(out_df.columns)}")


# ============================================================
# 미분류 Report
# ============================================================
def print_unclassified_report(df):
    print("\n" + "=" * 60)
    print("[미분류 Report]")
    print("=" * 60)

    checks = [
        ('channel_aware_rms', 'channel_aware', '미분류', 'Channel 분류표', '인지 채널'),
        ('channel_apply_rms', 'channel_apply', '미분류', 'Channel 분류표', '지원 채널'),
        ('channel_reuse_rms', 'channel_reuse', '미분류', 'Channel 분류표', '재사용 채널'),
        ('seg_income', 'income', '', 'Segment 분류표 (소득)', 'Seg 소득'),
    ]

    has_issues = False
    for result_col, raw_col, bad_val, table_name, label in checks:
        mask = (df[result_col] == bad_val) & (df[raw_col] != '')
        problem_df = df[mask]
        if len(problem_df) == 0:
            continue
        has_issues = True
        grouped = problem_df.groupby(raw_col).apply(
            lambda g: g.index.tolist(), include_groups=False)
        print(f"\n  {label}: {len(problem_df)}건 미분류 → {table_name}에 추가 필요")
        for raw_val, indices in grouped.items():
            row_nums = [str(idx + 2) for idx in indices[:5]]
            row_str = ', '.join(row_nums)
            if len(indices) > 5:
                row_str += f' 외 {len(indices)-5}건'
            print(f"    - \"{raw_val}\" ({len(indices)}건, 결과 행: {row_str})")

    if not has_issues:
        print("  미분류 없음!")


# ============================================================
# Main
# ============================================================
def main():
    import argparse
    parser = argparse.ArgumentParser(description='AM Placement Survey - R_통합 생성')
    parser.add_argument('--base', type=str, required=True,
                        help='이전 R_통합 결과 파일 또는 기존 분석 Excel')
    parser.add_argument('--raw', type=str, help='신규 분기 Raw 파일 경로')
    parser.add_argument('--quarter', type=str, help='신규 분기 이름 (예: 26.1Q)')
    parser.add_argument('--output', type=str, help='출력 파일 경로')
    parser.add_argument('--init', action='store_true',
                        help='기존 분석 Excel에서 초기 R_통합 생성 (1회용)')
    args = parser.parse_args()

    print("=" * 60)
    print("AM Placement Survey - R_통합 생성")
    print("=" * 60)

    # Step 1: 분류표
    print(f"\n[Step 1] Loading: {CLASSIFICATION_FILE.name}")
    cls = load_classifications(CLASSIFICATION_FILE)
    print(f"  Channel: {len(cls['channel'])} | Reason: {len(cls['reason'])}")
    print(f"  Cubicle age rules: {len(cls['cubicle_age_rules'])} | region rules: {len(cls['cubicle_region_rules'])}")

    # Step 2: Base data
    base_path = Path(args.base)
    if not base_path.exists():
        print(f"  ERROR: {base_path} not found")
        return

    if args.init:
        print(f"\n[Step 2] Reading historical from: {base_path.name}")
        df = read_historical_raw(base_path)
    else:
        print(f"\n[Step 2] Reading base: {base_path.name}")
        df = read_base_data(base_path)
    print(f"  Quarters: {sorted(df['quarter'].unique().tolist())}")

    # Step 3: New raw
    if args.raw and args.quarter:
        raw_path = Path(args.raw)
        if not raw_path.exists():
            print(f"  ERROR: {raw_path} not found")
            return
        print(f"\n[Step 3] Parsing {args.quarter}...")
        new_df = parse_raw_file(raw_path, args.quarter)
        df = df[df['quarter'] != args.quarter]
        df = pd.concat([df, new_df], ignore_index=True)
        print(f"  Total: {len(df)} rows")

    # Step 4: Classify
    print("\n[Step 4] Classifying...")
    df = apply_all_classifications(df, cls)

    # 미분류 리포트
    print_unclassified_report(df)

    # Step 5: Output
    if args.output:
        output_path = Path(args.output)
    elif args.quarter:
        output_path = AM_DIR / f"{args.quarter}_AM_결과.xlsx"
    else:
        output_path = AM_DIR / "AM_결과.xlsx"

    print(f"\n[Step 5] Writing output...")
    write_output(df, output_path)
    print("\nDone!")


if __name__ == '__main__':
    main()
