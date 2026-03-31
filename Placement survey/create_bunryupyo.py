# -*- coding: utf-8 -*-
"""분류표.xlsx 생성: 기존 분석 Excel에서 5개 분류표 + Cubicle 규칙 추출"""
import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SRC = "JK 전체/25Q4_JK 잡플레이스먼트 분석_v0.xlsx"
OUT = "분류표.xlsx"

src_wb = openpyxl.load_workbook(SRC, data_only=True)
dst_wb = openpyxl.Workbook()
dst_wb.remove(dst_wb.active)

header_font = Font(bold=True, size=11)
header_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)


def style_header(ws, row, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')


# ============================================================
# 1. Channel 분류표
# ============================================================
ws_src = src_wb['Channel 분류표']
ws_dst = dst_wb.create_sheet('Channel 분류표')
ws_dst.append(['채널명', 'ON/OFF', 'PAID/UNPAID', '채널 Segment', '채널 재분류'])
style_header(ws_dst, 1, 5)

for row in ws_src.iter_rows(min_row=2, min_col=6, max_col=10):
    key = row[0].value
    if key and str(key).strip():
        ws_dst.append([
            str(row[0].value).strip() if row[0].value else '',
            str(row[1].value).strip() if row[1].value else '',
            str(row[2].value).strip() if row[2].value else '',
            str(row[3].value).strip() if row[3].value else '',
            str(row[4].value).strip() if row[4].value else '',
        ])

ws_dst.column_dimensions['A'].width = 45
ws_dst.column_dimensions['E'].width = 15
print(f"Channel 분류표: {ws_dst.max_row - 1} entries")

# ============================================================
# 2. 산업 분류표 (Cubicle용)
# ============================================================
ws_src = src_wb['산업 분류표']
ws_dst = dst_wb.create_sheet('산업 분류표')
ws_dst.append(['산업 응답값', '산업 Segment (9개)'])
style_header(ws_dst, 1, 2)

for row in ws_src.iter_rows(min_row=3, min_col=6, max_col=7):
    if row[0].value and row[1].value:
        ws_dst.append([str(row[0].value).strip(), str(row[1].value).strip()])

ws_dst.column_dimensions['A'].width = 30
ws_dst.column_dimensions['B'].width = 20
print(f"산업 분류표: {ws_dst.max_row - 1} entries")

# ============================================================
# 3. 지역 분류표
# ============================================================
ws_src = src_wb['지역 분류표']
ws_dst = dst_wb.create_sheet('지역 분류표')
ws_dst.append(['지역 응답값', '일반 지역 Segment', '', '지역 응답값(공공)', '공공 지역 Segment'])
style_header(ws_dst, 1, 5)

for row in ws_src.iter_rows(min_row=3, min_col=6, max_col=10):
    gk = str(row[0].value).strip() if row[0].value else ''
    gv = str(row[1].value).strip() if row[1].value else ''
    pk = str(row[3].value).strip() if row[3].value else ''
    pv = str(row[4].value).strip() if row[4].value else ''
    if gk or pk:
        ws_dst.append([gk, gv, '', pk, pv])

ws_dst.column_dimensions['A'].width = 20
ws_dst.column_dimensions['B'].width = 20
ws_dst.column_dimensions['D'].width = 20
ws_dst.column_dimensions['E'].width = 20
print(f"지역 분류표: {ws_dst.max_row - 1} entries")

# ============================================================
# 4. 이유 분류표
# ============================================================
ws_src = src_wb['이유 분류표']
ws_dst = dst_wb.create_sheet('이유 분류표')
ws_dst.append(['이유 텍스트', '이유 카테고리 (7개)'])
style_header(ws_dst, 1, 2)

for row in ws_src.iter_rows(min_row=1, min_col=6, max_col=7):
    if row[0].value and row[1].value:
        ws_dst.append([str(row[0].value).strip(), str(row[1].value).strip()])

ws_dst.column_dimensions['A'].width = 65
ws_dst.column_dimensions['B'].width = 20
print(f"이유 분류표: {ws_dst.max_row - 1} entries")

# ============================================================
# 5. 산업직무소득 Seg
# ============================================================
ws_src = src_wb['산업직무소득 seg']
ws_dst = dst_wb.create_sheet('산업직무소득 Seg')

ws_dst.append(['[산업 Seg]', '', '', '[직무 Seg]', '', '',
               '[소득 Seg]', '', '', '', '[지역 Seg]'])
ws_dst.append(['산업 응답값', '산업 Grouping', '', '직무 응답값', '직무 Grouping', '',
               'SQ11 연봉(연간)', 'DQ4 소득(월간)', '소득수준', '', '지역 응답값', '지역 Grouping'])
style_header(ws_dst, 2, 12)

for row in ws_src.iter_rows(min_row=3, min_col=1, max_col=17):
    out = [''] * 12
    if row[1].value and row[2].value:
        out[0] = str(row[1].value).strip()
        out[1] = str(row[2].value).strip()
    if row[4].value and row[5].value:
        out[3] = str(row[4].value).strip()
        out[4] = str(row[5].value).strip()
    if row[11].value:
        out[6] = str(row[11].value).strip()
    if row[12].value:
        out[7] = str(row[12].value).strip()
    if row[13].value:
        out[8] = str(row[13].value).strip()
    if row[15].value and row[16].value:
        out[10] = str(row[15].value).strip()
        out[11] = str(row[16].value).strip()
    if any(v for v in out):
        ws_dst.append(out)

for col, w in {'A': 25, 'B': 20, 'D': 35, 'E': 20, 'G': 30, 'H': 20, 'I': 10, 'K': 15, 'L': 15}.items():
    ws_dst.column_dimensions[col].width = w
print(f"산업직무소득 Seg: {ws_dst.max_row - 2} entries")

# ============================================================
# 6. Cubicle 규칙 (코드에서 분리 → Excel로 관리)
# ============================================================
ws_dst = dst_wb.create_sheet('Cubicle 규칙')
ws_dst.append(['산업 Group', '연령 Group 라벨', '포함 연령값 (쉼표 구분)', '지역 매핑'])
style_header(ws_dst, 1, 4)

cubicle_rules = [
    ('스킬기반산업', '2030', '20대,30대', 'general'),
    ('스킬기반산업', '4050', '40대,50대', '전국'),
    ('스킬기반산업', '60+', '60세 이상', '전국'),
    ('현장중심산업', '20~40', '20대,30대,40대', 'general'),
    ('현장중심산업', '50', '50대', '전국'),
    ('현장중심산업', '60+', '60세 이상', '전국'),
    ('전문직군산업', '20~50', '20대,30대,40대,50대', '전국'),
    ('전문직군산업', '60+', '60세 이상', '전국'),
    ('자격중심산업', '20~50', '20대,30대,40대,50대', '전국'),
    ('자격중심산업', '60+', '60세 이상', '전국'),
    ('공공형산업', '20~40', '20대,30대,40대', 'public'),
    ('공공형산업', '50', '50대', '전국'),
    ('공공형산업', '60+', '60세 이상', '전국'),
    ('프로젝트형산업', '2030', '20대,30대', '전국'),
    ('프로젝트형산업', '40', '40대', '전국'),
    ('프로젝트형산업', '50', '50대', '전국'),
    ('프로젝트형산업', '60+', '60세 이상', '전국'),
    ('대량채용산업', '20~40', '20대,30대,40대', 'general'),
    ('대량채용산업', '50', '50대', '전국'),
    ('대량채용산업', '60+', '60세 이상', '전국'),
    ('상시채용산업', '2030', '20대,30대', '전국'),
    ('상시채용산업', '4050', '40대,50대', '전국'),
    ('상시채용산업', '60+', '60세 이상', '전국'),
    ('기타산업', '기타', '20대,30대,40대,50대,60세 이상', '전국'),
]

for rule in cubicle_rules:
    ws_dst.append(list(rule))

r = ws_dst.max_row + 2
ws_dst.cell(row=r, column=1, value='[지역 매핑 설명]').font = Font(bold=True)
ws_dst.cell(row=r + 1, column=1, value='general = 지역 분류표의 "일반 지역 Segment" 사용')
ws_dst.cell(row=r + 2, column=1, value='public = 지역 분류표의 "공공 지역 Segment" 사용')
ws_dst.cell(row=r + 3, column=1, value='전국 = 지역 분할 없이 "전국"으로 통일')

ws_dst.column_dimensions['A'].width = 20
ws_dst.column_dimensions['B'].width = 18
ws_dst.column_dimensions['C'].width = 35
ws_dst.column_dimensions['D'].width = 15
print(f"Cubicle 규칙: {len(cubicle_rules)} rules")

# ============================================================
# Save
# ============================================================
dst_wb.save(OUT)
print(f"\n분류표.xlsx 생성 완료")
src_wb.close()
