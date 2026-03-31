# -*- coding: utf-8 -*-
"""AM 분류표 추출: 기존 분석 Excel → 25Q4_AM_분류표.xlsx"""
import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SRC = "AM 전체/25Q4_AM 잡플레이스먼트 분석_v0.xlsx"
OUT = "AM 전체/25Q4_AM_분류표.xlsx"

src_wb = openpyxl.load_workbook(SRC, data_only=True)
dst_wb = openpyxl.Workbook()
dst_wb.remove(dst_wb.active)

hf = Font(bold=True, size=11)
hfill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
bdr = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))

def sh(ws, r, max_c):
    for c in range(1, max_c + 1):
        cell = ws.cell(row=r, column=c)
        cell.font = hf; cell.fill = hfill; cell.border = bdr

# ============================================================
# 1. Channel 분류표
# ============================================================
ws_src = src_wb['Channel 분류표']
ws_dst = dst_wb.create_sheet('Channel 분류표')
ws_dst.append(['채널명', 'ON/OFF', 'PAID/UNPAID', '채널 재분류'])
sh(ws_dst, 1, 4)

cnt = 0
# Source: B=ON/OFF, C=PAID/UNPAID, D=raw channel, E=채널재정리
for row in ws_src.iter_rows(min_row=5, min_col=2, max_col=5):
    on_off = row[0].value   # B
    paid = row[1].value     # C
    raw = row[2].value      # D = raw channel
    target = row[3].value   # E = 채널재정리
    if raw and target and str(target).strip():
        ws_dst.append([
            str(raw).strip(),            # A = 채널명
            str(on_off).strip() if on_off else '',  # B = ON/OFF
            str(paid).strip() if paid else '',       # C = PAID/UNPAID
            str(target).strip(),         # D = 채널 재분류
        ])
        cnt += 1

ws_dst.column_dimensions['A'].width = 45
ws_dst.column_dimensions['D'].width = 18
print(f"Channel 분류표: {cnt} entries")

# ============================================================
# 2. Segment 분류표 (직무, 근무형태, 소득, 지역)
# ============================================================
ws_src = src_wb['Segment 분류표']
ws_dst = dst_wb.create_sheet('Segment 분류표')
ws_dst.append(['[직무 Seg]', '', '', '[근무형태 Seg]', '', '',
               '[소득 Seg]', '', '', '', '[지역 Seg]'])
ws_dst.append(['직무 응답값', '직무 Grouping', '', '계약기간 응답값', '근무형태 Grouping', '',
               '소득 응답값', '소득 Grouping', '', '', '지역 응답값', '지역 Grouping'])
sh(ws_dst, 2, 12)

# Read all data from Segment 분류표
seg_rows = []
for row in ws_src.iter_rows(min_row=3, min_col=1, max_col=17):
    out = [''] * 12
    # B→C: 직무
    if row[1].value and row[2].value:
        out[0] = str(row[1].value).strip()
        out[1] = str(row[2].value).strip()
    # E→F: 근무형태(계약기간)
    if row[4].value and row[5].value:
        out[3] = str(row[4].value).strip()
        out[4] = str(row[5].value).strip()
    # H→I: 소득 (쉼표 구분 복합값 → 개별 행으로 분리는 아래서 처리)
    if row[7].value and row[8].value:
        out[6] = str(row[7].value).strip()
        out[7] = str(row[8].value).strip()
    # P→Q: 지역
    if row[15].value and row[16].value:
        out[10] = str(row[15].value).strip()
        out[11] = str(row[16].value).strip()
    if any(v for v in out):
        ws_dst.append(out)
        seg_rows.append(out)

# 소득 매핑: 쉼표 구분 복합값을 개별 행으로 추가
income_extra = []
for sr in seg_rows:
    if sr[6] and ',' in sr[6]:
        parts = [p.strip() for p in sr[6].split(',')]
        for part in parts:
            if part:
                income_extra.append([part, sr[7]])
for key, val in income_extra:
    ws_dst.append(['', '', '', '', '', '', key, val, '', '', '', ''])

# "800만원 이상" 추가 (Raw에 존재하나 매핑 없음)
ws_dst.append(['', '', '', '', '', '', '800만원 이상', '고소득', '', '', '', ''])

for col, w in {'A': 30, 'B': 20, 'D': 30, 'E': 18, 'G': 30, 'H': 15, 'K': 18, 'L': 15}.items():
    ws_dst.column_dimensions[col].width = w
print(f"Segment 분류표: {len(seg_rows)} rows (+{len(income_extra)+1} income splits)")

# ============================================================
# 3. 이유 분류표
# ============================================================
ws_src = src_wb['지원 채널 선택 이유 분류']
ws_dst = dst_wb.create_sheet('이유 분류표')
ws_dst.append(['이유 텍스트', '이유 카테고리'])
sh(ws_dst, 1, 2)

cnt = 0
for row in ws_src.iter_rows(min_row=2, min_col=3, max_col=4):
    key, val = row[0].value, row[1].value
    if key and val:
        ws_dst.append([str(key).strip(), str(val).strip()])
        cnt += 1

ws_dst.column_dimensions['A'].width = 65
ws_dst.column_dimensions['B'].width = 25
print(f"이유 분류표: {cnt} entries")

# ============================================================
# 4. Cubicle 규칙
# ============================================================
ws_src = src_wb['Cubicle_vF']
ws_dst = dst_wb.create_sheet('Cubicle 규칙')
ws_dst.append(['성별', '연령 Group', '거주지 (LV3)', '거주지 재분류 (LV3 merged)', '최종 Cubicle 라벨'])
sh(ws_dst, 1, 5)

cnt = 0
for row in ws_src.iter_rows(min_row=4, max_row=33, min_col=1, max_col=5):
    gender = row[0].value
    age_grp = row[1].value
    region = row[2].value
    region2 = row[3].value
    label = row[4].value
    if gender or age_grp or region:
        ws_dst.append([
            str(gender).strip() if gender else '',
            str(age_grp).strip() if age_grp else '',
            str(region).strip() if region else '',
            str(region2).strip() if region2 else '',
            str(label).strip() if label else '',
        ])
        cnt += 1

ws_dst.column_dimensions['A'].width = 8
ws_dst.column_dimensions['B'].width = 18
ws_dst.column_dimensions['C'].width = 15
ws_dst.column_dimensions['D'].width = 25
ws_dst.column_dimensions['E'].width = 30
print(f"Cubicle 규칙: {cnt} entries")

# ============================================================
# 5. Cubicle 연령 규칙 (코드용 - 성별+연령대 → 연령Group)
# ============================================================
ws_dst2 = dst_wb.create_sheet('Cubicle 연령 규칙')
ws_dst2.append(['성별', '연령대 응답값 (쉼표 구분)', '연령 Group'])
sh(ws_dst2, 1, 3)

age_rules = [
    ('남자', '16-19세,40세 이상', '10대+40대 이상'),
    ('남자', '20-24세,25-29세', '20대'),
    ('남자', '30-39세', '30대'),
    ('여자', '16-19세,30-39세', '10대+30대'),
    ('여자', '20-24세,25-29세', '20대'),
    ('여자', '40세 이상', '40대 이상'),
]
for rule in age_rules:
    ws_dst2.append(list(rule))

ws_dst2.column_dimensions['A'].width = 8
ws_dst2.column_dimensions['B'].width = 35
ws_dst2.column_dimensions['C'].width = 18
print(f"Cubicle 연령 규칙: {len(age_rules)} rules")

# ============================================================
# Save
# ============================================================
dst_wb.save(OUT)
print(f"\n25Q4_AM_분류표.xlsx 생성 완료")
src_wb.close()
