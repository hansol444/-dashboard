import openpyxl
import re
from collections import defaultdict

TARGET_SHEETS = [
    '1. 지원 RMS', '1-1. 채널 Cut', '1-2. Cubicle Cut', '1-3. Seg Cut',
    '1-4. Seg 시사점', '1-5. 지원 채널 선택 이유',
    '2. 인지 RMS', '2-1. 채널 Cut',
    '3. 재지원 RMS', '3-1. 채널 Cut', '3-2. 재지원 채널 선택 이유',
    'Cover'
]

def get_col_letter(col_num):
    result = ""
    while col_num > 0:
        col_num, remainder = divmod(col_num - 1, 26)
        result = chr(65 + remainder) + result
    return result

def cell_ref(row, col):
    return f"{get_col_letter(col)}{row}"

def parse_args_toplevel(inner):
    """Parse top-level comma-separated args respecting nesting."""
    parts = []
    depth = 0
    current = ""
    for ch in inner:
        if ch == '(':
            depth += 1
            current += ch
        elif ch == ')':
            depth -= 1
            current += ch
        elif ch == ',' and depth == 0:
            parts.append(current.strip())
            current = ""
        else:
            current += ch
    if current.strip():
        parts.append(current.strip())
    return parts

def extract_col_from_range(rng):
    m = re.findall(r'\$?([A-Z]+)\$?\d+:\$?([A-Z]+)\$?\d+', rng, re.IGNORECASE)
    if m:
        return m[0][0].upper()
    return None

def extract_row_range(rng):
    m = re.findall(r'\$?[A-Z]+\$?(\d+):\$?[A-Z]+\$?(\d+)', rng, re.IGNORECASE)
    if m:
        return (int(m[0][0]), int(m[0][1]))
    return None

def extract_balanced(formula, start_pos):
    """Extract balanced parenthesized content starting after opening paren."""
    depth = 1
    end = start_pos
    while end < len(formula) and depth > 0:
        if formula[end] == '(':
            depth += 1
        elif formula[end] == ')':
            depth -= 1
        end += 1
    return formula[start_pos:end-1], end

def audit_file(filepath, label):
    print(f"\n{'='*80}")
    print(f"AUDITING: {label}")
    print(f"{'='*80}")

    wb = openpyxl.load_workbook(filepath, data_only=False)
    bugs = []
    formula_count = 0
    sheet_formula_counts = {}

    available_sheets = [s for s in wb.sheetnames if s in TARGET_SHEETS]
    print(f"Target sheets found: {len(available_sheets)}")

    for ws_name in available_sheets:
        ws = wb[ws_name]
        sheet_formulas = 0
        formulas_by_type = defaultdict(int)

        for row_cells in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row_cells:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    formula = cell.value
                    formula_count += 1
                    sheet_formulas += 1
                    r = cell.row
                    c = cell.column
                    loc = cell_ref(r, c)
                    f_upper = formula.upper()

                    # === CHECK: COUNTIFS column issues ===
                    if 'COUNTIFS(' in f_upper:
                        formulas_by_type['COUNTIFS'] += 1
                        idx = 0
                        while True:
                            pos = f_upper.find('COUNTIFS(', idx)
                            if pos == -1:
                                break
                            start = pos + len('COUNTIFS(')
                            inner, end = extract_balanced(formula, start)
                            idx = end

                            parts = parse_args_toplevel(inner)

                            # Check odd args
                            if len(parts) % 2 != 0:
                                bugs.append({
                                    'sheet': ws_name, 'cell': loc, 'formula': formula,
                                    'issue': f'COUNTIFS has odd number of arguments ({len(parts)})',
                                    'severity': 'HIGH'
                                })
                                continue

                            range_args = [parts[i] for i in range(0, len(parts), 2)]
                            range_cols = []
                            range_rows = []
                            for rng in range_args:
                                col_l = extract_col_from_range(rng)
                                row_r = extract_row_range(rng)
                                if col_l:
                                    range_cols.append(col_l)
                                if row_r:
                                    range_rows.append(row_r)

                            # All ranges same column with 3+ criteria pairs
                            if len(range_cols) >= 3:
                                if len(set(range_cols)) == 1:
                                    bugs.append({
                                        'sheet': ws_name, 'cell': loc, 'formula': formula,
                                        'issue': f'COUNTIFS: ALL {len(range_cols)} criteria ranges use column {range_cols[0]} - copy-paste bug',
                                        'severity': 'CRITICAL'
                                    })

                            # Inconsistent row ranges
                            if len(range_rows) >= 2:
                                sizes = [(e - s + 1) for s, e in range_rows]
                                if len(set(sizes)) > 1:
                                    bugs.append({
                                        'sheet': ws_name, 'cell': loc, 'formula': formula,
                                        'issue': f'COUNTIFS: Mismatched range sizes: {dict(zip(range_cols, [f"rows {s}-{e} (size {e-s+1})" for s,e in range_rows]))}',
                                        'severity': 'MEDIUM'
                                    })

                    # === CHECK: SUMPRODUCT array size mismatch ===
                    elif 'SUMPRODUCT(' in f_upper:
                        formulas_by_type['SUMPRODUCT'] += 1
                        idx = 0
                        while True:
                            pos = f_upper.find('SUMPRODUCT(', idx)
                            if pos == -1:
                                break
                            start = pos + len('SUMPRODUCT(')
                            inner, end = extract_balanced(formula, start)
                            idx = end

                            row_ranges = re.findall(r'\$?[A-Z]+\$?(\d+):\$?[A-Z]+\$?(\d+)', inner)
                            if len(row_ranges) >= 2:
                                sizes = [(int(e) - int(s) + 1) for s, e in row_ranges]
                                if len(set(sizes)) > 1:
                                    detail = [(s, e, int(e)-int(s)+1) for s, e in row_ranges]
                                    bugs.append({
                                        'sheet': ws_name, 'cell': loc, 'formula': formula,
                                        'issue': f'SUMPRODUCT: Array SIZE mismatch - {detail}',
                                        'severity': 'HIGH'
                                    })

                    elif '/' in formula:
                        formulas_by_type['Division'] += 1
                    else:
                        formulas_by_type['Other'] += 1

                    # === CHECK: #REF! ===
                    if '#REF!' in formula:
                        bugs.append({
                            'sheet': ws_name, 'cell': loc, 'formula': formula,
                            'issue': 'Broken reference #REF!',
                            'severity': 'CRITICAL'
                        })

                    # === CHECK: Mismatched parentheses ===
                    open_p = formula.count('(')
                    close_p = formula.count(')')
                    if open_p != close_p:
                        bugs.append({
                            'sheet': ws_name, 'cell': loc, 'formula': formula,
                            'issue': f'Mismatched parentheses: {open_p}( vs {close_p})',
                            'severity': 'HIGH'
                        })

                    # === CHECK: Non-existent sheet reference ===
                    sheet_refs = re.findall(r"'([^']+)'!", formula)
                    for sr in sheet_refs:
                        if sr not in wb.sheetnames and '#REF' not in sr:
                            bugs.append({
                                'sheet': ws_name, 'cell': loc, 'formula': formula,
                                'issue': f'References non-existent sheet: "{sr}"',
                                'severity': 'CRITICAL'
                            })

                    # === CHECK: Self-reference ===
                    col_letter = get_col_letter(c)
                    # Only flag if not part of a range or sheet-qualified reference
                    # Simple heuristic: check for exact cell address
                    exact_addr = f"{col_letter}{r}"
                    exact_abs = f"${col_letter}${r}"
                    # Must not be preceded by : (range) or ! (sheet ref) or letter
                    pat1 = re.compile(r'(?<![A-Z:!])' + re.escape(exact_addr) + r'(?!\d)', re.IGNORECASE)
                    pat2 = re.compile(r'(?<![:!])' + re.escape(exact_abs) + r'(?!\d)', re.IGNORECASE)
                    formula_body = formula[1:]  # skip the =
                    if pat1.search(formula_body) or pat2.search(formula_body):
                        # Exclude if it's inside a sheet!ref pattern
                        # Check that it's not preceded by '!' within a few chars
                        for pat in [pat1, pat2]:
                            m = pat.search(formula_body)
                            if m:
                                before = formula_body[max(0,m.start()-1):m.start()]
                                if before != '!':
                                    bugs.append({
                                        'sheet': ws_name, 'cell': loc, 'formula': formula,
                                        'issue': f'CIRCULAR: Cell references itself ({exact_addr})',
                                        'severity': 'HIGH'
                                    })
                                    break

        sheet_formula_counts[ws_name] = (sheet_formulas, dict(formulas_by_type))

    # Print summary
    print(f"\nTotal formulas in target sheets: {formula_count}")
    for sn, (cnt, types) in sheet_formula_counts.items():
        if cnt > 0:
            print(f"  {sn}: {cnt} ({types})")

    # Deduplicate
    unique_bugs = []
    seen = set()
    for bug in bugs:
        norm = re.sub(r'\d+', 'N', bug['formula'][:80])
        key = (bug['sheet'], bug['issue'][:40], norm)
        if key not in seen:
            seen.add(key)
            bug['instances'] = 1
            bug['other_cells'] = []
            unique_bugs.append(bug)
        else:
            for ub in unique_bugs:
                norm2 = re.sub(r'\d+', 'N', ub['formula'][:80])
                key2 = (ub['sheet'], ub['issue'][:40], norm2)
                if key2 == key:
                    ub['instances'] += 1
                    if len(ub['other_cells']) < 30:
                        ub['other_cells'].append(bug['cell'])
                    break

    print(f"\n{'='*80}")
    print(f"BUGS: {len(unique_bugs)} unique patterns ({len(bugs)} total instances)")
    print(f"{'='*80}")

    for i, bug in enumerate(unique_bugs, 1):
        f_display = bug['formula'][:500]
        n = bug['instances']
        print(f"\n--- Bug #{i} [{bug['severity']}] x{n} ---")
        print(f"  Sheet: {bug['sheet']}")
        print(f"  Cell:  {bug['cell']}" + (f" + {n-1} more" if n > 1 else ""))
        if bug['other_cells']:
            print(f"  Also: {', '.join(bug['other_cells'][:15])}")
        print(f"  Formula: {f_display}")
        print(f"  Issue: {bug['issue']}")

    return bugs, unique_bugs


print("=" * 80)
print("COMPREHENSIVE FORMULA AUDIT v3")
print("=" * 80)

jk_all, jk_unique = audit_file(
    r"c:\Users\ugin35\Desktop\Placement survey 자동화 revive\JK 전체\25Q4_JK 잡플레이스먼트 분석_v0.xlsx",
    "JK File"
)

am_all, am_unique = audit_file(
    r"c:\Users\ugin35\Desktop\Placement survey 자동화 revive\AM 전체\25Q4_AM 잡플레이스먼트 분석_v0.xlsx",
    "AM File"
)

print(f"\n\n{'#'*80}")
print(f"GRAND TOTAL: JK={len(jk_all)} instances/{len(jk_unique)} patterns, AM={len(am_all)} instances/{len(am_unique)} patterns")
