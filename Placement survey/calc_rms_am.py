# -*- coding: utf-8 -*-
"""
AM Placement Survey - RMS 계산 + 결과 시트 생성
Usage: python calc_rms_am.py --input "AM 전체/25Q4_AM_결과.xlsx"
"""
import sys
sys.stdout.reconfigure(encoding='utf-8')

import argparse
import numpy as np
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side

# ============================================================
# Constants (AM)
# ============================================================
CHANNELS = ['알바몬', '알바천국', '당근(알바)', '온라인 Others', '오프라인 Paid', '오프라인 Unpaid', 'Out']
CHANNELS_ONLINE = ['알바몬', '알바천국', '당근(알바)', '온라인 Others']
CHANNELS_4 = ['알바몬', '알바천국', '당근(알바)']  # AM 3사
REASON_CATS = ['아르바이트 공고', '유용한 정보 제공', '이용 편의성', '부가서비스 및 혜택',
               '사용자 경험 및 인지도', '위치 및 채용 특성', '그외 (기타)']

CUBICLE_ORDER = [
    '남자10대+40대 이상경기/강원권', '남자10대+40대 이상서울', '남자10대+40대 이상영남권',
    '남자10대+40대 이상충청권', '남자10대+40대 이상호남권',
    '남자20대경기/강원권 + 서울', '남자20대영남권', '남자20대충청권', '남자20대호남권',
    '남자30대경기/강원권', '남자30대서울', '남자30대영남권', '남자30대충청권', '남자30대호남권',
    '여자10대+30대경기/강원권', '여자10대+30대서울', '여자10대+30대영남권',
    '여자10대+30대충청권', '여자10대+30대호남권',
    '여자20대경기/강원권', '여자20대서울', '여자20대영남권', '여자20대충청권', '여자20대호남권',
    '여자40대 이상경기/강원권 + 서울', '여자40대 이상영남권', '여자40대 이상충청권', '여자40대 이상호남권',
]

# AM 분류표 경로 (Population weight 로드용)
AM_CLASSIFICATION = Path(r"C:\Users\ugin35\Desktop\Placement survey 자동화 revive\AM 전체\25Q4_AM_분류표.xlsx")

SCOPES = [
    ('전체 (Online+Offline)', CHANNELS),
    ('ONLINE only', CHANNELS_ONLINE),
    ('AM AH DG', CHANNELS_4),  # 알바몬/알바천국/당근(알바) 3사
]

SEG_DIMS = [
    ('seg_job', '직무'), ('seg_contract', '근무형태'),
    ('seg_income', '소득'), ('seg_region', '지역'),
]

# Styling
HDR_FONT = Font(bold=True, size=10)
HDR_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
SEC_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
BORDER = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
PCT = '0.0%'


def sc(ws, r, c, v, fmt=None, bold=False, fill=None):
    """Style-write a cell"""
    cell = ws.cell(row=r, column=c, value=v)
    if fmt:
        cell.number_format = fmt
    if bold:
        cell.font = HDR_FONT
    if fill:
        cell.fill = fill
    cell.border = BORDER
    return cell


# ============================================================
# Data Loading
# ============================================================
def load_population_weight(cls_path=AM_CLASSIFICATION):
    """Population Weight 시트에서 고정 인구비례 가중치 로드"""
    import openpyxl as ox
    wb = ox.load_workbook(cls_path, data_only=True)
    ws = wb['Population Weight']
    pop_wt = {}
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=2):
        cub = str(row[0].value).strip() if row[0].value else ''
        wt = row[1].value
        if cub and wt:
            pop_wt[cub] = float(wt)
    wb.close()
    return pop_wt


def load_data(filepath):
    df = pd.read_excel(filepath, sheet_name='R_통합', engine='openpyxl')
    rename = {
        'Quarter': 'quarter', '최종 Cubicle': 'cubicle',
        '지원 RMS 재분류': 'ch_apply', '인지 RMS 재분류': 'ch_aware',
        '재사용 RMS 재분류': 'ch_reuse',
        'Why 지원 분류 1': 'ra1', 'Why 지원 분류 2': 'ra2', 'Why 지원 분류 3': 'ra3',
        'Why 재지원 분류 1': 'rr1', 'Why 재지원 분류 2': 'rr2', 'Why 재지원 분류 3': 'rr3',
        'Seg 직무': 'seg_job', 'Seg 근무형태': 'seg_contract',
        'Seg 소득': 'seg_income', 'Seg 지역': 'seg_region',
    }
    df = df.rename(columns=rename).fillna('')
    for c in df.columns:
        if df[c].dtype == object:
            df[c] = df[c].astype(str).str.strip().replace('nan', '')
    return df


# ============================================================
# Core Calculations
# ============================================================
def calc_rms(df, ch_col, quarters):
    """Cubicle-based RMS: unadjusted, weight, adjusted, totals"""
    unadj = {ch: {q: {} for q in quarters} for ch in CHANNELS}
    weight = {q: {} for q in quarters}

    for q in quarters:
        qdf = df[df['quarter'] == q]
        nq = len(qdf)
        if nq == 0:
            continue
        for cub in CUBICLE_ORDER:
            nc = len(qdf[qdf['cubicle'] == cub])
            weight[q][cub] = nc / nq
            cub_df = qdf[qdf['cubicle'] == cub]
            for ch in CHANNELS:
                unadj[ch][q][cub] = len(cub_df[cub_df[ch_col] == ch]) / nc if nc > 0 and len(cub_df) > 0 else 0
                # Fix: should be / len(cub_df)
    # Recalculate properly
    unadj = {ch: {q: {} for q in quarters} for ch in CHANNELS}
    for q in quarters:
        qdf = df[df['quarter'] == q]
        nq = len(qdf)
        for cub in CUBICLE_ORDER:
            cub_df = qdf[qdf['cubicle'] == cub]
            nc = len(cub_df)
            for ch in CHANNELS:
                if nc > 0:
                    unadj[ch][q][cub] = len(cub_df[cub_df[ch_col] == ch]) / nc
                else:
                    unadj[ch][q][cub] = 0

    # AM Adjusted: unadjusted * Population Weight (고정 인구비례)
    pop_wt = load_population_weight()
    adj = {ch: {q: {} for q in quarters} for ch in CHANNELS}
    for ch in CHANNELS:
        for q in quarters:
            for cub in CUBICLE_ORDER:
                adj[ch][q][cub] = unadj[ch][q].get(cub, 0) * pop_wt.get(cub, 0)

    # Totals
    t_adj = {ch: {} for ch in CHANNELS}
    t_unadj = {ch: {} for ch in CHANNELS}
    for ch in CHANNELS:
        for q in quarters:
            t_adj[ch][q] = sum(adj[ch][q].get(c, 0) for c in CUBICLE_ORDER)
            t_unadj[ch][q] = sum(
                unadj[ch][q].get(c, 0) * weight[q].get(c, 0) for c in CUBICLE_ORDER)

    return {'unadj_cub': unadj, 'weight': weight, 'adj_cub': adj,
            't_adj': t_adj, 't_unadj': t_unadj}


def calc_reasons(df, ch_col, rcols, quarters, ch_filter=None):
    """3-2-1 weighted reason scoring"""
    w = [3, 2, 1]
    scores, shares = {}, {}
    for q in quarters:
        qdf = df[df['quarter'] == q]
        if ch_filter:
            qdf = qdf[qdf[ch_col] == ch_filter]
        scores[q] = {}
        for cat in REASON_CATS:
            scores[q][cat] = sum(len(qdf[qdf[rc] == cat]) * w[i] for i, rc in enumerate(rcols))
        tot = sum(scores[q].values())
        shares[q] = {cat: scores[q][cat] / tot if tot else 0 for cat in REASON_CATS}
    return scores, shares


def calc_seg(df, ch_col, seg_col, quarters, seg_vals):
    """Seg Cut: channel share per segment per quarter"""
    res = {}
    for seg in seg_vals:
        res[seg] = {ch: {} for ch in CHANNELS}
        for q in quarters:
            sdf = df[(df['quarter'] == q) & (df[seg_col] == seg)]
            n = len(sdf)
            for ch in CHANNELS:
                res[seg][ch][q] = len(sdf[sdf[ch_col] == ch]) / n if n > 0 else 0
    return res


def get_seg_values(df, seg_col):
    """Seg 고유값 추출 (빈값 제외, 기타 마지막)"""
    vals = sorted(set(v for v in df[seg_col].unique() if v and v not in ('', 'nan', '미분류')))
    if '기타' in vals:
        vals.remove('기타')
        vals.append('기타')
    return vals


# ============================================================
# Sheet: RMS Summary (1 per type, Adj+Unadj+3scopes)
# ============================================================
def write_rms_sheet(ws, name, rms, quarters):
    ta, tu = rms['t_adj'], rms['t_unadj']
    nq = len(quarters)

    for adj_label, data, start_row in [('Adjusted RMS', ta, 2), ('Unadjusted RMS', tu, None)]:
        if start_row is None:
            start_row = ws.max_row + 3

        sc(ws, start_row, 1, f'{name} {adj_label}', bold=True)
        r = start_row + 1

        for scope_name, scope_channels in SCOPES:
            # Section header
            sc(ws, r, 1, scope_name, bold=True, fill=SEC_FILL)
            for i, q in enumerate(quarters):
                sc(ws, r, i + 3, q, bold=True, fill=HDR_FILL)
            sc(ws, r, nq + 3, 'QoQ', bold=True, fill=HDR_FILL)
            r += 1

            # Channel rows
            for ch in scope_channels:
                sc(ws, r, 2, ch)
                for i, q in enumerate(quarters):
                    sc(ws, r, i + 3, data[ch].get(q, 0), fmt=PCT)
                if nq >= 2:
                    curr = data[ch].get(quarters[-1], 0)
                    prev = data[ch].get(quarters[-2], 0)
                    sc(ws, r, nq + 3, curr - prev, fmt=PCT)
                r += 1

            # 총계
            sc(ws, r, 2, '총계', bold=True)
            for i, q in enumerate(quarters):
                t = sum(data[ch].get(q, 0) for ch in scope_channels)
                sc(ws, r, i + 3, t, fmt=PCT)
            r += 1

            # Share (각 채널 / scope 총계)
            sc(ws, r, 1, f'{scope_name} Share', bold=True, fill=SEC_FILL)
            r += 1
            for ch in scope_channels:
                sc(ws, r, 2, ch)
                for i, q in enumerate(quarters):
                    t = sum(data[c].get(q, 0) for c in scope_channels)
                    sc(ws, r, i + 3, data[ch].get(q, 0) / t if t else 0, fmt=PCT)
                if nq >= 2:
                    tc = sum(data[c].get(quarters[-1], 0) for c in scope_channels)
                    tp = sum(data[c].get(quarters[-2], 0) for c in scope_channels)
                    cc = data[ch].get(quarters[-1], 0) / tc if tc else 0
                    cp = data[ch].get(quarters[-2], 0) / tp if tp else 0
                    sc(ws, r, nq + 3, cc - cp, fmt=PCT)
                r += 1
            r += 1

        # AM/(AM+AH)
        sc(ws, r, 1, 'AM/(AM+AH)', bold=True)
        for i, q in enumerate(quarters):
            am = data['알바몬'].get(q, 0)
            ah = data['알바천국'].get(q, 0)
            sc(ws, r, i + 3, am / (am + ah) if (am + ah) else 0, fmt=PCT)
        r += 1

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 14


# ============================================================
# Sheet: 채널 Cut (cubicle-level detail per channel)
# ============================================================
def write_channel_cut(ws, name, rms, quarters):
    unadj = rms['unadj_cub']
    adj_c = rms['adj_cub']
    wt = rms['weight']
    nq = len(quarters)
    col_start = 1

    for ch in CHANNELS:
        r = 1
        sc(ws, r, col_start, f'{ch}', bold=True, fill=HDR_FILL)
        sc(ws, r, col_start + 1, 'Cubicle', bold=True, fill=HDR_FILL)
        for i, q in enumerate(quarters):
            sc(ws, r, col_start + 2 + i, q, bold=True, fill=HDR_FILL)
        r += 1

        # Total Adjusted
        sc(ws, r, col_start, 'Total Adj', bold=True)
        for i, q in enumerate(quarters):
            sc(ws, r, col_start + 2 + i,
               sum(adj_c[ch][q].get(c, 0) for c in CUBICLE_ORDER), fmt=PCT)
        r += 1

        # Total Unadjusted
        sc(ws, r, col_start, 'Total Unadj', bold=True)
        for i, q in enumerate(quarters):
            sc(ws, r, col_start + 2 + i,
               sum(unadj[ch][q].get(c, 0) * wt[q].get(c, 0) for c in CUBICLE_ORDER), fmt=PCT)
        r += 1

        # Cubicle Adjusted RMS
        sc(ws, r, col_start, 'Adj RMS', bold=True, fill=SEC_FILL)
        r += 1
        for cub in CUBICLE_ORDER:
            sc(ws, r, col_start + 1, cub)
            for i, q in enumerate(quarters):
                sc(ws, r, col_start + 2 + i, adj_c[ch][q].get(cub, 0), fmt=PCT)
            r += 1

        # Cubicle Unadjusted RMS
        sc(ws, r, col_start, 'Unadj RMS', bold=True, fill=SEC_FILL)
        r += 1
        for cub in CUBICLE_ORDER:
            sc(ws, r, col_start + 1, cub)
            for i, q in enumerate(quarters):
                sc(ws, r, col_start + 2 + i, unadj[ch][q].get(cub, 0), fmt=PCT)
            r += 1

        # Weight
        sc(ws, r, col_start, 'Weight', bold=True, fill=SEC_FILL)
        r += 1
        for cub in CUBICLE_ORDER:
            sc(ws, r, col_start + 1, cub)
            for i, q in enumerate(quarters):
                sc(ws, r, col_start + 2 + i, wt[q].get(cub, 0), fmt=PCT)
            r += 1

        col_start += nq + 3  # next channel block

    ws.column_dimensions['B'].width = 35


# ============================================================
# Sheet: Seg Cut
# ============================================================
def write_seg_cut(ws, name, df, ch_col, quarters):
    """Seg Cut: 산업/직무/소득/지역을 가로로 배치 (옆으로 이동해서 볼 수 있도록)"""
    nq = len(quarters)
    block_width = nq + 3  # seg_name + channel + quarters + gap

    sc(ws, 1, 1, f'{name} Seg Cut', bold=True)

    for dim_idx, (seg_col, seg_name) in enumerate(SEG_DIMS):
        col_start = dim_idx * block_width + 1
        seg_vals = get_seg_values(df, seg_col)
        seg_data = calc_seg(df, ch_col, seg_col, quarters, seg_vals)

        r = 3
        sc(ws, r, col_start, f'[{seg_name}]', bold=True, fill=SEC_FILL)
        r += 1

        # Header
        sc(ws, r, col_start, seg_name, bold=True, fill=HDR_FILL)
        sc(ws, r, col_start + 1, 'Channel', bold=True, fill=HDR_FILL)
        for i, q in enumerate(quarters):
            sc(ws, r, col_start + 2 + i, q, bold=True, fill=HDR_FILL)
        r += 1

        for seg in seg_vals:
            sc(ws, r, col_start, seg, bold=True)
            sc(ws, r, col_start + 1, '(N)')
            for i, q in enumerate(quarters):
                qdf = df[df['quarter'] == q]
                sc(ws, r, col_start + 2 + i, len(qdf[qdf[seg_col] == seg]))
            r += 1

            # 4사 Share
            for ch in CHANNELS_4:
                sc(ws, r, col_start + 1, ch)
                for i, q in enumerate(quarters):
                    four = sum(seg_data[seg][c].get(q, 0) for c in CHANNELS_4)
                    sc(ws, r, col_start + 2 + i,
                       seg_data[seg][ch].get(q, 0) / four if four else 0, fmt=PCT)
                r += 1

            # 전체 채널
            for ch in CHANNELS:
                sc(ws, r, col_start + 1, ch)
                for i, q in enumerate(quarters):
                    sc(ws, r, col_start + 2 + i, seg_data[seg][ch].get(q, 0), fmt=PCT)
                r += 1
            r += 1


# ============================================================
# Sheet: Seg 그래프 (1년/3년 평균)
# ============================================================
def write_seg_graph(ws, name, df, ch_col, quarters):
    """Seg 그래프: 1년/3년 평균, 가로 배치 (산업→직무→소득→지역 옆으로)"""
    nq = len(quarters)
    q_1yr = quarters[-4:] if nq >= 4 else quarters
    q_3yr = quarters[-12:] if nq >= 12 else quarters
    block_width = 6  # seg + channel + 1yr + 3yr + diff + gap

    sc(ws, 1, 1, f'{name} Seg 그래프 데이터', bold=True)
    sc(ws, 2, 1, f'1년: {q_1yr[0]}~{q_1yr[-1]} ({len(q_1yr)}Q) / 3년: {q_3yr[0]}~{q_3yr[-1]} ({len(q_3yr)}Q)')

    for dim_idx, (seg_col, seg_name) in enumerate(SEG_DIMS):
        seg_vals = get_seg_values(df, seg_col)
        col_s = dim_idx * block_width + 1

        r = 4
        sc(ws, r, col_s, f'[{seg_name}]', bold=True, fill=SEC_FILL)
        r += 1
        sc(ws, r, col_s, seg_name, bold=True, fill=HDR_FILL)
        sc(ws, r, col_s + 1, 'Channel', bold=True, fill=HDR_FILL)
        sc(ws, r, col_s + 2, '1년 평균', bold=True, fill=HDR_FILL)
        sc(ws, r, col_s + 3, '3년 평균', bold=True, fill=HDR_FILL)
        sc(ws, r, col_s + 4, '차이', bold=True, fill=HDR_FILL)
        r += 1

        for seg in seg_vals:
            sc(ws, r, col_s, seg, bold=True)
            sc(ws, r, col_s + 1, '(N 1yr)')
            n1 = sum(len(df[(df['quarter'] == q) & (df[seg_col] == seg)]) for q in q_1yr)
            sc(ws, r, col_s + 2, n1)
            r += 1

            for ch in CHANNELS_4:
                sc(ws, r, col_s + 1, ch)
                vals_1, vals_3 = [], []
                for q in q_1yr:
                    sdf = df[(df['quarter'] == q) & (df[seg_col] == seg)]
                    if len(sdf) > 0:
                        four = sum(len(sdf[sdf[ch_col] == c]) for c in CHANNELS_4)
                        vals_1.append(len(sdf[sdf[ch_col] == ch]) / four if four else 0)
                for q in q_3yr:
                    sdf = df[(df['quarter'] == q) & (df[seg_col] == seg)]
                    if len(sdf) > 0:
                        four = sum(len(sdf[sdf[ch_col] == c]) for c in CHANNELS_4)
                        vals_3.append(len(sdf[sdf[ch_col] == ch]) / four if four else 0)
                a1 = np.mean(vals_1) if vals_1 else 0
                a3 = np.mean(vals_3) if vals_3 else 0
                sc(ws, r, col_s + 2, a1, fmt=PCT)
                sc(ws, r, col_s + 3, a3, fmt=PCT)
                sc(ws, r, col_s + 4, a1 - a3, fmt=PCT)
                r += 1
            r += 1


# ============================================================
# Sheet: 채널 선택 이유
# ============================================================
def write_reason_sheet(ws, title, df, ch_col, rcols, quarters, exclude_quarters=None):
    """채널 선택 이유. exclude_quarters: 제외할 분기 리스트"""
    if exclude_quarters:
        quarters = [q for q in quarters if q not in exclude_quarters]
    sc(ws, 1, 1, title, bold=True)
    sc(ws, 2, 1, '가중치: 1순위×3, 2순위×2, 3순위×1')
    nq = len(quarters)

    blocks = [(None, '전체'), ('알바몬', '알바몬'), ('알바천국', '알바천국'),
              ('당근(알바)', '당근(알바)')]
    r = 4
    for ch_f, ch_name in blocks:
        sc(ws, r, 1, ch_name, bold=True, fill=HDR_FILL)
        sc(ws, r, 2, '이유 분류', bold=True, fill=HDR_FILL)
        col = 3
        for q in quarters:
            sc(ws, r, col, f'{q} 점수', bold=True, fill=HDR_FILL)
            sc(ws, r, col + 1, f'{q} 비중', bold=True, fill=HDR_FILL)
            col += 2
        r += 1

        scores, shares = calc_reasons(df, ch_col, rcols, quarters, ch_filter=ch_f)
        for cat in REASON_CATS:
            sc(ws, r, 2, cat)
            col = 3
            for q in quarters:
                sc(ws, r, col, scores.get(q, {}).get(cat, 0))
                sc(ws, r, col + 1, shares.get(q, {}).get(cat, 0), fmt=PCT)
                col += 2
            r += 1
        r += 1

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20


def write_seg_reason_sheet(ws, title, df, ch_col, rcols, quarters, exclude_quarters=None):
    """Seg×채널 이유 비중 시트. 각 Seg dimension별, 각 Seg값별, 각 채널별 이유 Top3 비중"""
    if exclude_quarters:
        quarters = [q for q in quarters if q not in exclude_quarters]

    sc(ws, 1, 1, title, bold=True)
    sc(ws, 2, 1, '가중치: 1순위×3, 2순위×2, 3순위×1 / 최근 4Q 합산')
    r = 4

    # 최근 4Q만 사용 (sample size)
    recent_q = [quarters[-1]]  # 최신 1Q만

    for seg_col, seg_name in SEG_DIMS:
        sc(ws, r, 1, f'[{seg_name}]', bold=True, fill=SEC_FILL)
        r += 1

        # Header
        sc(ws, r, 1, 'Seg', bold=True, fill=HDR_FILL)
        sc(ws, r, 2, 'Channel', bold=True, fill=HDR_FILL)
        sc(ws, r, 3, 'N', bold=True, fill=HDR_FILL)
        for i, cat in enumerate(REASON_CATS):
            sc(ws, r, 4 + i, cat, bold=True, fill=HDR_FILL)
        r += 1

        seg_vals = get_seg_values(df, seg_col)
        for seg_val in seg_vals:
            sdf = df[(df[seg_col] == seg_val) & (df['quarter'].isin(recent_q))]
            if len(sdf) == 0:
                continue

            channels_to_check = CHANNELS_4  # 3사만
            first_ch = True
            for ch in channels_to_check:
                chdf = sdf[sdf[ch_col] == ch]
                n = len(chdf)
                if n == 0:
                    continue

                # 3-2-1 weighted scoring
                w = [3, 2, 1]
                cat_scores = {}
                for cat in REASON_CATS:
                    cat_scores[cat] = sum(
                        len(chdf[chdf[rc] == cat]) * w[i]
                        for i, rc in enumerate(rcols))
                tot = sum(cat_scores.values())
                cat_shares = {cat: cat_scores[cat] / tot if tot else 0
                              for cat in REASON_CATS}

                if first_ch:
                    sc(ws, r, 1, seg_val)
                    first_ch = False
                sc(ws, r, 2, ch)
                sc(ws, r, 3, n)
                for i, cat in enumerate(REASON_CATS):
                    sc(ws, r, 4 + i, cat_shares[cat], fmt=PCT)
                r += 1
            r += 1  # 빈 행

        r += 1  # Seg dimension 구분

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 6


# ============================================================
# Sheet: 지원 Seg 시사점 (데이터 + 그룹분류 + 시사점 템플릿)
# ============================================================
def write_seg_insight(ws, df, ch_col, quarters):
    """Seg 시사점 시트: 소득수준/산업/직무 각각에 대해
    - 과거(1yr avg) vs 현재(latest Q) 비교 테이블
    - 온라인 PTR + 표본비중 + 4사 RMS
    - Group 분류 (PTR 기준)
    - JK-SRI delta → 우위/on-par/열위
    - 시사점 작성 영역 (빈 칸)
    """
    latest_q = quarters[-1]
    nq = len(quarters)
    q_past = quarters[-5:-1] if nq >= 5 else quarters[:-1]  # 최신 제외 직전 4Q

    r = 1
    sc(ws, r, 1, '지원 Seg 시사점', bold=True)
    r += 1
    sc(ws, r, 1, f'현재: {latest_q} / 과거: {q_past[0]}~{q_past[-1]} ({len(q_past)}Q 평균)' if q_past else f'현재: {latest_q}')
    r += 2

    # 범례
    sc(ws, r, 1, '[JK vs SRI 판정 기준]', bold=True)
    r += 1
    sc(ws, r, 1, '+%p'); sc(ws, r, 2, '우위 (JK > SRI)')
    r += 1
    sc(ws, r, 1, '±10% 이내'); sc(ws, r, 2, 'on-par')
    r += 1
    sc(ws, r, 1, '-%p'); sc(ws, r, 2, '열위 (JK < SRI)')
    r += 2

    insight_dims = [
        ('seg_income', '소득수준'),
        ('seg_ind', '산업'),
        ('seg_job', '직무'),
    ]

    for seg_col, seg_name in insight_dims:
        seg_vals = get_seg_values(df, seg_col)

        sc(ws, r, 1, f'[{seg_name} Cut]', bold=True, fill=SEC_FILL)
        r += 1

        # ---- 데이터 테이블 ----
        # 과거 RMS 블록
        past_label = f'과거 ({q_past[0]}~{q_past[-1]})' if q_past else '과거'
        curr_label = f'현재 ({latest_q})'

        # 헤더
        headers = ['Group', '온라인 PTR', '표본비중', seg_name,
                   'JK RMS', 'SRI RMS', 'RMB RMS', 'WTD RMS', 'JK-SRI', '판정']
        # 과거 블록
        sc(ws, r, 1, past_label, bold=True, fill=HDR_FILL)
        for i, h in enumerate(headers):
            sc(ws, r, i + 1, h, bold=True, fill=HDR_FILL)
        # 현재 블록 (옆에)
        off = len(headers) + 1
        sc(ws, r, off, curr_label, bold=True, fill=HDR_FILL)
        for i, h in enumerate(headers):
            sc(ws, r, off + i, h, bold=True, fill=HDR_FILL)
        r += 1

        # 각 seg에 대해 데이터 계산
        seg_rows = []
        for seg in seg_vals:
            # --- 과거 평균 ---
            past_online_ptrs = []
            past_sample_pcts = []
            past_rms = {ch: [] for ch in CHANNELS_4}

            for q in q_past:
                qdf = df[df['quarter'] == q]
                sdf = qdf[qdf[seg_col] == seg]
                n_q = len(qdf)
                n_s = len(sdf)
                if n_s == 0:
                    continue

                # 온라인 PTR = 온라인 채널 합 / 전체
                online_n = sum(len(sdf[sdf[ch_col] == c]) for c in CHANNELS_ONLINE)
                past_online_ptrs.append(online_n / n_s)
                past_sample_pcts.append(n_s / n_q)

                # 4사 RMS
                four_n = sum(len(sdf[sdf[ch_col] == c]) for c in CHANNELS_4)
                for ch in CHANNELS_4:
                    ch_n = len(sdf[sdf[ch_col] == ch])
                    past_rms[ch].append(ch_n / four_n if four_n else 0)

            p_ptr = np.mean(past_online_ptrs) if past_online_ptrs else 0
            p_sample = np.mean(past_sample_pcts) if past_sample_pcts else 0
            p_rms_vals = {ch: np.mean(past_rms[ch]) if past_rms[ch] else 0 for ch in CHANNELS_4}
            p_jk_sri = p_rms_vals['알바몬'] - p_rms_vals['알바천국']

            # --- 현재 ---
            qdf = df[df['quarter'] == latest_q]
            sdf = qdf[qdf[seg_col] == seg]
            n_q = len(qdf)
            n_s = len(sdf)

            if n_s > 0:
                online_n = sum(len(sdf[sdf[ch_col] == c]) for c in CHANNELS_ONLINE)
                c_ptr = online_n / n_s
                c_sample = n_s / n_q
                four_n = sum(len(sdf[sdf[ch_col] == c]) for c in CHANNELS_4)
                c_rms = {}
                for ch in CHANNELS_4:
                    ch_n = len(sdf[sdf[ch_col] == ch])
                    c_rms[ch] = ch_n / four_n if four_n else 0
                c_jk_sri = c_rms['알바몬'] - c_rms['알바천국']
            else:
                c_ptr = 0
                c_sample = 0
                c_rms = {ch: 0 for ch in CHANNELS_4}
                c_jk_sri = 0

            # Group 분류 (현재 PTR 기준)
            if c_sample < 0.005:
                group = 'Group 4. 표본 부족'
            elif c_ptr >= 0.25:
                group = 'Group 1. PTR HIGH'
            elif c_ptr >= 0.15:
                group = 'Group 2. PTR MID'
            else:
                group = 'Group 3. PTR LOW'

            # 판정
            def judge(delta):
                if abs(delta) <= 0.10:
                    return 'on-par'
                return '우위' if delta > 0 else '열위'

            seg_rows.append({
                'seg': seg, 'group': group,
                'p_ptr': p_ptr, 'p_sample': p_sample, 'p_rms': p_rms_vals,
                'p_delta': p_jk_sri, 'p_judge': judge(p_jk_sri),
                'c_ptr': c_ptr, 'c_sample': c_sample, 'c_rms': c_rms,
                'c_delta': c_jk_sri, 'c_judge': judge(c_jk_sri),
            })

        # Group 순으로 정렬
        group_order = {'Group 1. PTR HIGH': 0, 'Group 2. PTR MID': 1,
                       'Group 3. PTR LOW': 2, 'Group 4. 표본 부족': 3}
        seg_rows.sort(key=lambda x: (group_order.get(x['group'], 9), -x['c_ptr']))

        # 데이터 출력
        current_group = ''
        for sr in seg_rows:
            grp_label = sr['group'] if sr['group'] != current_group else ''
            current_group = sr['group']

            # 과거 블록
            sc(ws, r, 1, grp_label, bold=bool(grp_label))
            sc(ws, r, 2, sr['p_ptr'], fmt=PCT)
            sc(ws, r, 3, sr['p_sample'], fmt=PCT)
            sc(ws, r, 4, sr['seg'])
            for i, ch in enumerate(CHANNELS_4):
                sc(ws, r, 5 + i, sr['p_rms'][ch], fmt=PCT)
            sc(ws, r, 9, sr['p_delta'], fmt='+0.0%;-0.0%')
            sc(ws, r, 10, sr['p_judge'])

            # 현재 블록
            sc(ws, r, off, sr['group'] if grp_label else '')
            sc(ws, r, off + 1, sr['c_ptr'], fmt=PCT)
            sc(ws, r, off + 2, sr['c_sample'], fmt=PCT)
            sc(ws, r, off + 3, sr['seg'])
            for i, ch in enumerate(CHANNELS_4):
                sc(ws, r, off + 4 + i, sr['c_rms'][ch], fmt=PCT)
            sc(ws, r, off + 8, sr['c_delta'], fmt='+0.0%;-0.0%')
            sc(ws, r, off + 9, sr['c_judge'])
            r += 1

        r += 1

        # ---- 시사점 템플릿 ----
        sc(ws, r, 1, '1) Overall Trend', bold=True, fill=SEC_FILL)
        for i in range(5):
            sc(ws, r + 1 + i, 2, '')  # 빈 줄 5개
        r += 7

        sc(ws, r, 1, '2) Segment Deep-dive', bold=True, fill=SEC_FILL)
        r += 1
        for grp_name in ['Group 1.', 'Group 2.', 'Group 3.', 'Group 4.']:
            sc(ws, r, 1, grp_name, bold=True)
            grp_segs = [sr for sr in seg_rows if sr['group'].startswith(grp_name.replace('.', ''))]
            for sr in grp_segs:
                r += 1
                delta_str = f"{sr['c_delta']*100:+.1f}%p"
                sc(ws, r, 2, delta_str)
                sc(ws, r, 3, f"- ({sr['seg']}) ")  # 시사점 작성 영역
            r += 2
        r += 2

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 22
    for c_idx in range(5, 11):
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(c_idx)].width = 10


# ============================================================
# Main
# ============================================================
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--input', required=True)
    parser.add_argument('--output', default=None)
    args = parser.parse_args()

    inp = Path(args.input)
    out = Path(args.output) if args.output else inp.parent / (inp.stem.replace('_결과', '') + '_RMS.xlsx')

    print("=" * 50)
    print("JK RMS 계산")
    print("=" * 50)

    print(f"\n[1] Reading: {inp.name}")
    df = load_data(inp)
    quarters = sorted(df['quarter'].unique().tolist())
    print(f"  {len(df)} rows, {len(quarters)} quarters: {quarters[0]}~{quarters[-1]}")

    # RMS types
    types = [('지원', 'ch_apply', ['ra1', 'ra2', 'ra3']),
             ('인지', 'ch_aware', None),
             ('재지원', 'ch_reuse', ['rr1', 'rr2', 'rr3'])]

    rms_results = {}
    for name, col, _ in types:
        print(f"[2] {name} RMS 계산...")
        rms_results[name] = calc_rms(df, col, quarters)
        am = rms_results[name]['t_adj']['알바몬'].get(quarters[-1], 0)
        print(f"  {quarters[-1]} AM Adj={am:.4f}")

    print(f"\n[3] Writing: {out.name}")
    wb = Workbook()
    wb.remove(wb.active)

    for name, col, rcols in types:
        # RMS Summary (1시트: Adj+Unadj+3scopes)
        ws = wb.create_sheet(f'{name} RMS')
        write_rms_sheet(ws, name, rms_results[name], quarters)

        # 채널 Cut
        ws = wb.create_sheet(f'{name} 채널 Cut')
        write_channel_cut(ws, name, rms_results[name], quarters)

        # Seg Cut
        ws = wb.create_sheet(f'{name} Seg Cut')
        write_seg_cut(ws, name, df, col, quarters)

        # Seg 그래프
        ws = wb.create_sheet(f'{name} Seg 그래프')
        write_seg_graph(ws, name, df, col, quarters)

    # 채널 선택 이유
    # AM 지원: 25.1Q부터 이유 데이터 수집
    apply_exclude = [q for q in quarters if q < '25.1Q']
    ws = wb.create_sheet('지원 채널 선택 이유')
    write_reason_sheet(ws, '지원 채널 선택 이유', df, 'ch_apply',
                       ['ra1', 'ra2', 'ra3'], quarters, exclude_quarters=apply_exclude)

    # AM 재지원: 25.3Q부터 이유 데이터 수집
    reuse_exclude = [q for q in quarters if q < '25.3Q']
    ws = wb.create_sheet('재지원 채널 선택 이유')
    write_reason_sheet(ws, '재지원 채널 선택 이유', df, 'ch_reuse',
                       ['rr1', 'rr2', 'rr3'], quarters, exclude_quarters=reuse_exclude)

    # Seg×채널 이유 비중
    ws = wb.create_sheet('지원 Seg×채널 이유')
    write_seg_reason_sheet(ws, '지원 Seg×채널 이유 비중', df, 'ch_apply',
                           ['ra1', 'ra2', 'ra3'], quarters, exclude_quarters=apply_exclude)

    ws = wb.create_sheet('재지원 Seg×채널 이유')
    write_seg_reason_sheet(ws, '재지원 Seg×채널 이유 비중', df, 'ch_reuse',
                           ['rr1', 'rr2', 'rr3'], quarters, exclude_quarters=reuse_exclude)

    wb.save(out)
    print(f"  Sheets: {wb.sheetnames}")
    print(f"  Total: {len(wb.sheetnames)} sheets")
    print("\nDone!")


if __name__ == '__main__':
    main()
