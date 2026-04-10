import openpyxl
import re
import sys
from collections import defaultdict

def get_col_letter(col_num):
    result = ""
    while col_num > 0:
        col_num, remainder = divmod(col_num - 1, 26)
        result = chr(65 + remainder) + result
    return result

def cell_ref(row, col):
    return f"{get_col_letter(col)}{row}"

def audit_file(filepath, label):
    print(f"\n{'='*80}")
    print(f"AUDITING: {label}")
    print(f"File: {filepath}")
    print(f"{'='*80}")

    wb = openpyxl.load_workbook(filepath, data_only=False)

    bugs = []
    formula_count = 0
    sheet_formula_counts = {}

    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        sheet_formulas = 0

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    formula = cell.value
                    formula_count += 1
                    sheet_formulas += 1
                    r = cell.row
                    c = cell.column
                    loc = cell_ref(r, c)

                    # CHECK 1: Self-reference (circular)
                    cell_addr = f"{get_col_letter(c)}{r}"
                    cell_addr_abs = f"${get_col_letter(c)}${r}"
                    pattern_self = re.compile(r'(?<![A-Z])' + re.escape(cell_addr) + r'(?![0-9])', re.IGNORECASE)
                    pattern_self_abs = re.compile(re.escape(cell_addr_abs), re.IGNORECASE)
                    if pattern_self.search(formula) or pattern_self_abs.search(formula):
                        bugs.append({
                            'sheet': ws_name, 'cell': loc, 'formula': formula,
                            'issue': 'CIRCULAR REFERENCE - cell references itself',
                            'impact': 'Will cause circular reference error or incorrect calculation'
                        })

                    # CHECK 2: COUNTIFS column repetition bug
                    if 'COUNTIFS(' in formula.upper():
                        countifs_matches = re.findall(r'COUNTIFS\(([^)]+)\)', formula, re.IGNORECASE)
                        for cm in countifs_matches:
                            parts = []
                            depth = 0
                            current = ""
                            for ch in cm:
                                if ch == '(':
                                    depth += 1
                                    current += ch
                                elif ch == ')':
                                    depth -= 1
                                    current += ch
                                elif ch == ',' and depth == 0:
                                    parts.append(current.strip())
                                    current = ""
                                else:
                                    current += ch
                            if current.strip():
                                parts.append(current.strip())

                            # Even indices are ranges, odd are criteria
                            ranges = [parts[i] for i in range(0, len(parts), 2) if i < len(parts)]

                            range_cols = []
                            for rng in ranges:
                                col_match = re.findall(r'\$?([A-Z]+)\$?\d+:\$?([A-Z]+)\$?\d+', rng, re.IGNORECASE)
                                if col_match:
                                    range_cols.append(col_match[0][0].upper())

                            if len(range_cols) >= 3:
                                col_counts = defaultdict(int)
                                for rc in range_cols:
                                    col_counts[rc] += 1
                                for col_letter, count in col_counts.items():
                                    if count > 1 and count == len(range_cols):
                                        bugs.append({
                                            'sheet': ws_name, 'cell': loc, 'formula': formula,
                                            'issue': f'COUNTIFS: ALL {count} range arguments reference column {col_letter} - likely copy-paste bug',
                                            'impact': 'Counting with wrong criteria columns, producing incorrect counts'
                                        })

                            # Check odd number of args
                            if len(parts) % 2 != 0:
                                bugs.append({
                                    'sheet': ws_name, 'cell': loc, 'formula': formula,
                                    'issue': f'COUNTIFS has odd number of arguments ({len(parts)})',
                                    'impact': 'Formula error - should be pairs of range,criteria'
                                })

                    # CHECK 3: SUMPRODUCT array size mismatch
                    if 'SUMPRODUCT(' in formula.upper():
                        sp_matches = re.findall(r'SUMPRODUCT\(([^)]+)\)', formula, re.IGNORECASE)
                        for sm in sp_matches:
                            row_ranges = re.findall(r'\$?[A-Z]+\$?(\d+):\$?[A-Z]+\$?(\d+)', sm)
                            if len(row_ranges) >= 2:
                                sizes = set()
                                for start, end in row_ranges:
                                    sizes.add((int(start), int(end)))
                                if len(sizes) > 1:
                                    bugs.append({
                                        'sheet': ws_name, 'cell': loc, 'formula': formula,
                                        'issue': f'SUMPRODUCT: Array size mismatch - row ranges: {sizes}',
                                        'impact': 'Will cause #VALUE! error or incorrect calculation'
                                    })

                    # CHECK 4: Division by literal zero
                    if '/' in formula:
                        if re.search(r'/\s*0\s*[),+\-*]', formula) or formula.rstrip().endswith('/0'):
                            bugs.append({
                                'sheet': ws_name, 'cell': loc, 'formula': formula,
                                'issue': 'Division by literal zero',
                                'impact': '#DIV/0! error'
                            })

                    # CHECK 5: Suspicious large row numbers
                    large_rows = re.findall(r'\$?[A-Z]+\$?(\d{3,})', formula)
                    for lr in large_rows:
                        row_num = int(lr)
                        if row_num > 1000:
                            all_rows_in_formula = [int(x) for x in re.findall(r'\$?[A-Z]+\$?(\d+)', formula)]
                            if len(all_rows_in_formula) >= 3:
                                median_row = sorted(all_rows_in_formula)[len(all_rows_in_formula)//2]
                                if row_num > median_row * 5 and row_num > 500:
                                    bugs.append({
                                        'sheet': ws_name, 'cell': loc, 'formula': formula,
                                        'issue': f'Suspicious large row reference: row {row_num} (other refs near row {median_row})',
                                        'impact': 'Possibly a typo in row number, referencing wrong data'
                                    })

                    # CHECK 6: #REF! in formulas
                    if '#REF!' in formula:
                        bugs.append({
                            'sheet': ws_name, 'cell': loc, 'formula': formula,
                            'issue': 'Formula contains #REF! - broken reference',
                            'impact': 'Formula will produce #REF! error'
                        })

                    # CHECK 7: Mismatched parentheses
                    open_p = formula.count('(')
                    close_p = formula.count(')')
                    if open_p != close_p:
                        bugs.append({
                            'sheet': ws_name, 'cell': loc, 'formula': formula,
                            'issue': f'Mismatched parentheses: {open_p} open vs {close_p} close',
                            'impact': 'Formula syntax error'
                        })

                    # CHECK 8: Cross-sheet reference consistency
                    # Look for sheet references that might be wrong
                    sheet_refs = re.findall(r"'([^']+)'!", formula)
                    for sr in sheet_refs:
                        if sr not in wb.sheetnames and '#REF' not in sr:
                            bugs.append({
                                'sheet': ws_name, 'cell': loc, 'formula': formula,
                                'issue': f'References non-existent sheet: {sr}',
                                'impact': 'Formula will produce #REF! error'
                            })

        sheet_formula_counts[ws_name] = sheet_formulas

    print(f"\nTotal formulas found: {formula_count}")
    print(f"\nFormulas per sheet:")
    for sn, cnt in sheet_formula_counts.items():
        if cnt > 0:
            print(f"  {sn}: {cnt}")

    print(f"\n{'='*80}")
    print(f"BUGS FOUND: {len(bugs)}")
    print(f"{'='*80}")

    for i, bug in enumerate(bugs, 1):
        f_display = bug['formula'][:300]
        if len(bug['formula']) > 300:
            f_display += '...'
        print(f"\n--- Bug #{i} ---")
        print(f"  Sheet: {bug['sheet']}")
        print(f"  Cell:  {bug['cell']}")
        print(f"  Formula: {f_display}")
        print(f"  Issue: {bug['issue']}")
        print(f"  Impact: {bug['impact']}")

    return bugs

jk_bugs = audit_file(
    r"C:\Users\olivia408\projects\-dashboard\Placement survey\JK 전체\25Q4_JK 잡플레이스먼트 분석_v0.xlsx",
    "JK File"
)

am_bugs = audit_file(
    r"C:\Users\olivia408\projects\-dashboard\Placement survey\AM 전체\25Q4_AM 잡플레이스먼트 분석_v0.xlsx",
    "AM File"
)

print(f"\n\n{'#'*80}")
print(f"GRAND TOTAL: {len(jk_bugs) + len(am_bugs)} bugs found across both files")
print(f"  JK: {len(jk_bugs)} bugs")
print(f"  AM: {len(am_bugs)} bugs")
